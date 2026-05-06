#Versionfinal
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from dotenv import load_dotenv
from supabase import create_client
import os
from datetime import datetime, date, time as dt_time, timedelta
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import re
import unicodedata
import logging
import traceback
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()  # ✅ Carga el .env al iniciar

app = Flask(__name__)

# ✅ Permite peticiones desde Live Server (5500) y cualquier origen local
CORS(app, resources={r"/*": {"origins": "*"}})

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════
# CONEXIÓN SUPABASE
# ══════════════════════════════════════════════

def get_supabase():
    """Retorna una instancia del cliente Supabase."""
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_API_KEY")
    if not url or not key:
        raise RuntimeError("Faltan SUPABASE_URL o SUPABASE_API_KEY en el .env")
    return create_client(url, key)

def registrar_historial_subida(sb, pozo=None, no_instalacion=None, tipo=None, usuario=None, archivo=None, estado="OK", detalle=None):
    try:
        payload = {
            "pozo": pozo,
            "no_instalacion": str(no_instalacion) if no_instalacion is not None else None,
            "tipo": tipo,
            "usuario": usuario,
            "archivo": archivo,
            "estado": estado,
            "detalle": detalle,
        }

        sb.table("historial_subidas").insert(payload).execute()

    except Exception as e:
        logger.warning(f"No se pudo registrar historial: {e}")


@app.route("/api/historial-subidas", methods=["GET"])
def listar_historial_subidas():
    try:
        sb = get_supabase()

        res = (
            sb.table("historial_subidas")
            .select("fecha,pozo,no_instalacion,tipo,usuario,archivo,estado,detalle")
            .order("fecha", desc=True)
            .limit(50)
            .execute()
        )

        return jsonify({"ok": True, "data": serialize_rows(res.data)})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
def serialize_value(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return val


def serialize_row(row: dict) -> dict:
    return {k: serialize_value(v) for k, v in row.items()}


def serialize_rows(rows: list) -> list:
    return [serialize_row(r) for r in rows]


def _strip_accents(text: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(ch)
    )

def _normalize_text(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text == "" or text == "/" or text == "-":
        return None
    return text

def _normalize_header(text: str) -> str:
    text = _strip_accents(str(text or "")).upper().strip()
    text = re.sub(r"[\s\.\-\/]+", "_", text)
    text = re.sub(r"[^A-Z0-9_]", "", text)
    return text

def _normalize_column_name(text: str) -> str:
    """Normaliza nombres de columnas manteniendo MAYÚSCULAS para Supabase."""
    text = _strip_accents(str(text or "")).upper().strip()
    text = re.sub(r"[\s\.\-\/]+", "_", text)
    text = re.sub(r"[^A-Z0-9_]", "", text)
    return text

def _cell_value(ws, ref: str):
    """
    Devuelve el valor de la celda respetando merge areas.
    """
    cell = ws[ref]
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            return ws[merged.start_cell.coordinate].value
    return cell.value

def _normalize_cell(value):
    value = _normalize_text(value)
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return None if v == "" or v == "/" else v
    return value

def _split_by_slash(value):
    value = _normalize_cell(value)
    if not value or not isinstance(value, str):
        return value, None
    if "/" not in value:
        return value.strip(), None
    a, b = value.split("/", 1)
    return (a.strip() or None), (b.strip() or None)

def _split_name_parenthesis(value):
    value = _normalize_cell(value)
    if not value or not isinstance(value, str):
        return value, None
    m = re.match(r"^(.*?)\s*\((.*?)\)\s*$", value)
    if m:
        return (m.group(1).strip() or None), (m.group(2).strip() or None)
    return value.strip(), None

def _parse_long(value):
    value = _normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip().lower()
        if v in ["", "none", "null", "undefined", "nan"]:
            return None
        text = str(value).strip()
    if text == "":
        return None
    text = text.replace(" ", "")
    text = text.replace(".", "")
    text = text.replace(",", ".")
    try:
        return int(float(text))
    except Exception:
        digits = re.findall(r"-?\d+", text)
        return int(digits[0]) if digits else None

def _parse_float(value):
    value = _normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "":
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        digits = re.findall(r"-?\d+(?:\.\d+)?", text)
        return float(digits[0]) if digits else None

from datetime import datetime, date
import re

MESES = {
    "enero": "01", "febrero": "02", "marzo": "03",
    "abril": "04", "mayo": "05", "junio": "06",
    "julio": "07", "agosto": "08", "septiembre": "09",
    "octubre": "10", "noviembre": "11", "diciembre": "12"
}

def _to_supabase_value(value):
    value = _normalize_cell(value)

    if value is None:
        return None

    # 🔹 Si ya es datetime
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    # 🔹 Si es date
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    # 🔹 Si es string (caso Excel con meses en español)
    if isinstance(value, str):
        try:
            value = value.strip()

            # Detectar formato tipo: 27-enero-2026 13:00
            match = re.match(r"(\d{1,2})-([a-zA-Z]+)-(\d{4})\s+(\d{1,2}:\d{2})", value)

            if match:
                dia, mes_txt, anio, hora = match.groups()
                mes = MESES.get(mes_txt.lower())

                if mes:
                    return f"{anio}-{mes}-{int(dia):02d} {hora}:00"

            # Si viene solo fecha sin hora: 27-enero-2026
            match_fecha = re.match(r"(\d{1,2})-([a-zA-Z]+)-(\d{4})", value)

            if match_fecha:
                dia, mes_txt, anio = match_fecha.groups()
                mes = MESES.get(mes_txt.lower())

                if mes:
                    return f"{anio}-{mes}-{int(dia):02d}"

        except:
            return None

    return value

def _merge_top_left_text(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            return ws[merged.start_cell.coordinate].value
    return cell.value

def _concat_merged_row(ws, row: int, col_start: int, col_end: int, sep: str = " "):
    parts = []
    c = col_start
    while c <= col_end:
        cell = ws.cell(row=row, column=c)
        merged = None
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                merged = rng
                break

        if merged is not None:
            if merged.min_col == c:
                v = ws[merged.start_cell.coordinate].value
                v = _normalize_cell(v)
                if v not in (None, ""):
                    parts.append(str(v).strip())
            c = merged.max_col + 1
        else:
            v = _normalize_cell(cell.value)
            if v not in (None, ""):
                parts.append(str(v).strip())
            c += 1
    return sep.join(parts).strip() if parts else None


def _concat_range_raw(ws, row: int, start_col_letter: str, end_col_letter: str) -> str:
    """
    Equivalente a ConcatRangeRaw del VBA de Access:
    concatena el valor de cada bloque merged único en el rango de columnas,
    sin separador (igual que el VBA que usa res = res & CStr(c.Value)).
    """
    start_col = column_index_from_string(start_col_letter)
    end_col   = column_index_from_string(end_col_letter)
    result    = ""
    c = start_col
    while c <= end_col:
        cell   = ws.cell(row=row, column=c)
        merged = None
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                merged = rng
                break
        if merged is not None:
            if merged.min_col == c:
                val = ws[merged.start_cell.coordinate].value
                if val is not None:
                    result += str(val)
            c = merged.max_col + 1
        else:
            if cell.value is not None:
                result += str(cell.value)
            c += 1
    return result.strip()


def _combine_excel_date_time(raw_fecha, raw_hora) -> str | None:
    """
    Combina la parte de FECHA y la parte de HORA provenientes de celdas Excel
    (mismo patrón que el VBA para PARAMESTATIC_INSTALACION).
    Devuelve string "YYYY-MM-DD HH:MM:SS" o None.
    """
    # ── Parte FECHA ─────────────────────────────
    d_part = None
    if isinstance(raw_fecha, datetime):
        d_part = raw_fecha.date()
    elif isinstance(raw_fecha, date):
        d_part = raw_fecha
    elif isinstance(raw_fecha, (int, float)):
        try:
            d_part = (datetime(1899, 12, 30) + timedelta(days=float(raw_fecha))).date()
        except Exception:
            d_part = None
    elif isinstance(raw_fecha, str):
        s = raw_fecha.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                d_part = datetime.strptime(s, fmt).date()
                break
            except Exception:
                continue

    # ── Parte HORA ──────────────────────────────
    t_part = None
    if isinstance(raw_hora, dt_time):
        t_part = raw_hora
    elif isinstance(raw_hora, datetime):
        t_part = raw_hora.time()
    elif isinstance(raw_hora, (int, float)):
        try:
            frac = float(raw_hora) % 1.0
            total_s = int(round(frac * 86400))
            t_part  = dt_time(total_s // 3600, (total_s % 3600) // 60, total_s % 60)
        except Exception:
            t_part = None
    elif isinstance(raw_hora, str):
        s = raw_hora.strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                t_part = datetime.strptime(s, fmt).time()
                break
            except Exception:
                continue

    if d_part is None and t_part is None:
        return None
    if d_part is None:
        return str(t_part) if t_part else None
    if t_part is None:
        return d_part.strftime("%Y-%m-%d")
    return datetime.combine(d_part, t_part).strftime("%Y-%m-%d %H:%M:%S")


def _detect_header_cols(ws, header_row: int, max_col: int):
    """
    Detecta columnas por encabezado en una fila.
    Retorna dict con claves:
      EQUIPO, PARTE, SERIE, REEL, DESCRIP, CONDIC, PROPIED, OD, LONGITUD, PROF, SHIMS, GIRO, CANTIDAD, ACCESORIOS
    """
    cols = {
        "EQUIPO": 0, "PARTE": 0, "SERIE": 0, "REEL": 0, "DESCRIP": 0,
        "CONDIC": 0, "PROPIED": 0, "OD": 0, "LONGITUD": 0, "PROF": 0,
        "SHIMS": 0, "GIRO": 0, "CANTIDAD": 0, "ACCESORIOS": 0
    }

    for c in range(1, max_col + 1):
        hdr = _merge_top_left_text(ws, header_row, c)
        hdr = _normalize_header(hdr)
        if not hdr:
            continue

        if cols["EQUIPO"] == 0 and "EQUIPO" in hdr:
            cols["EQUIPO"] = c
        if cols["PARTE"] == 0 and "PARTE" in hdr:
            cols["PARTE"] = c
        if cols["SERIE"] == 0 and ("SERIE" in hdr or "NUMERO" in hdr or "SERIAL" in hdr):
            cols["SERIE"] = c
        if cols["REEL"] == 0 and "REEL" in hdr:
            cols["REEL"] = c
        if cols["DESCRIP"] == 0 and "DESCRIP" in hdr:
            cols["DESCRIP"] = c
        if cols["CONDIC"] == 0 and "CONDIC" in hdr:
            cols["CONDIC"] = c
        if cols["PROPIED"] == 0 and "PROPIED" in hdr:
            cols["PROPIED"] = c
        if cols["OD"] == 0 and (hdr == "OD" or "_OD" in hdr or hdr.endswith("OD")):
            cols["OD"] = c
        if cols["LONGITUD"] == 0 and "LONGITUD" in hdr:
            cols["LONGITUD"] = c
        if cols["PROF"] == 0 and "PROF" in hdr:
            cols["PROF"] = c
        if cols["SHIMS"] == 0 and "SHIMS" in hdr:
            cols["SHIMS"] = c
        if cols["GIRO"] == 0 and "GIRO" in hdr:
            cols["GIRO"] = c
        if cols["CANTIDAD"] == 0 and ("CANT" in hdr or "CANTIDAD" in hdr):
            cols["CANTIDAD"] = c
        if cols["ACCESORIOS"] == 0 and ("ACCESOR" in hdr or "EQUIPOS" in hdr):
            cols["ACCESORIOS"] = c

    return cols

def _json_safe(value):
    """Convierte objetos complejos a formatos seguros para JSON/logs."""
    if isinstance(value, (datetime, date, dt_time)):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(v) for v in value]
    return value

def _extract_exception_details(exc: Exception) -> dict:
    detail = {
        "type": type(exc).__name__,
        "message": str(exc),
    }
    if getattr(exc, "args", None):
        try:
            detail["args"] = _json_safe(exc.args)
        except Exception:
            detail["args"] = [str(a) for a in exc.args]
    return detail

def _append_warning(warnings: list, tabla: str, mensaje: str, **extra):
    item = {"tabla": tabla, "error": mensaje}
    for k, v in extra.items():
        item[k] = _json_safe(v)
    warnings.append(item)
    logger.warning("%s", item)

def _extract_missing_columns_from_error(err_detail) -> list[str]:
    """
    Extrae nombres de columnas ausentes desde mensajes de Supabase/PostgREST.
    """
    text = str(err_detail)
    missing = re.findall(r"Could not find the '([^']+)' column", text, flags=re.IGNORECASE)
    if missing:
        return list(dict.fromkeys(missing))
    missing = re.findall(r"column(?:s)? ['\"]([^'\"]+)['\"]", text, flags=re.IGNORECASE)
    return list(dict.fromkeys(missing))

def _to_int(value):
    try:
        if value is None:
            return None
        return int(float(value))
    except:
        return None
def _drop_keys_from_payload(payload: dict, keys_to_drop: list[str]) -> dict:
    if not keys_to_drop:
        return dict(payload)
    norm_drop = {_normalize_column_name(k) for k in keys_to_drop}
    out = {}
    for k, v in payload.items():
        if _normalize_column_name(k) in norm_drop:
            continue
        out[k] = v
    return out


def _normalize_payload_keys(payload: dict) -> dict:
    """Crea un payload con claves normalizadas para tablas definidas en minúsculas o snake_case."""
    return {_normalize_column_name(k): v for k, v in payload.items()}


def _safe_write(sb, table_name: str, payload: dict, *, mode: str = "insert", on_conflict: str | None = None, max_retries: int = 4):
    """
    Escritura tolerante a errores:
    - intenta insertar/upsert
    - si Supabase reporta una columna inexistente, la elimina y reintenta
    - si falla por otra causa, devuelve el detalle sin detener la importación
    """
    original_payload = dict(payload)
    current_payload = dict(payload)
    dropped_due_error = []
    last_error = None

    for _attempt in range(1, max_retries + 1):
        try:
            if mode == "upsert":
                query = sb.table(table_name).upsert(current_payload)
                if on_conflict:
                    try:
                        query = query.on_conflict(on_conflict)
                    except Exception:
                        pass
            else:
                query = sb.table(table_name).insert(current_payload)

            res = query.execute()
            return True, {
                "data": res.data if hasattr(res, "data") else None,
                "payload_original": _json_safe(original_payload),
                "payload_final": _json_safe(current_payload),
                "dropped_due_error": dropped_due_error,
                "mode": mode,
            }

        except Exception as e:
            last_error = _extract_exception_details(e)
            missing_cols = _extract_missing_columns_from_error(last_error)
            if missing_cols:
                new_payload = _drop_keys_from_payload(current_payload, missing_cols)
                if new_payload != current_payload:
                    dropped_due_error.extend(missing_cols)
                    current_payload = new_payload
                    continue
            break

    return False, {
        "error": "Falló la escritura en Supabase",
        "payload_original": _json_safe(original_payload),
        "payload_final": _json_safe(current_payload),
        "dropped_due_error": dropped_due_error,
        "detalle": _json_safe(last_error),
        "mode": mode,
    }


def _safe_insert(sb, table_name: str, payload: dict):
    return _safe_write(sb, table_name, payload, mode="insert")


def _safe_upsert(sb, table_name: str, payload: dict, on_conflict: str | None = None):
    return _safe_write(sb, table_name, payload, mode="upsert", on_conflict=on_conflict)


def _safe_write_with_variants(sb, table_name: str, payload: dict, *, mode: str = "insert", on_conflict: str | None = None):
    """Intenta escribir primero con las claves originales y luego con claves normalizadas."""
    attempts = [
        ("original", dict(payload)),
    ]
    normalized = _normalize_payload_keys(payload)
    if normalized != payload:
        attempts.append(("normalizado", normalized))

    last_result = None
    for variant_name, variant_payload in attempts:
        ok, result = _safe_write(sb, table_name, variant_payload, mode=mode, on_conflict=on_conflict)
        result = dict(result) if isinstance(result, dict) else {"resultado": result}
        result["variant"] = variant_name
        if ok:
            return True, result
        last_result = result

    return False, last_result or {"error": "Falló la escritura", "variant": None}

def _parse_int(value):
    value = _normalize_cell(value)
    if value is None:
        return None
    try:
        return int(round(float(value)))
    except Exception:
        try:
            return int(round(_parse_float(value)))
        except Exception:
            return None

def _extract_textbook_rows(ws, row_start: int, row_end: int, col_start: int, col_end: int):
    """
    Detecta bloques lógicos merged dentro de una fila y devuelve las columnas iniciales.
    """
    starts = []
    c = col_start
    while c <= col_end and len(starts) < 6:
        cell = ws.cell(row=row_start, column=c)
        merged = None
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                merged = rng
                break

        if merged is not None:
            if merged.min_col == c:
                starts.append(c)
                c = merged.max_col + 1
            else:
                c += 1
        else:
            starts.append(c)
            c += 1

    if len(starts) < 6:
        total_cols = col_end - col_start + 1
        width = max(1, total_cols // 6)
        starts = [col_start + i * width for i in range(6)]
    return starts


# ══════════════════════════════════════════════
# HEALTH CHECK
# ══════════════════════════════════════════════

@app.route("/api/health", methods=["GET"])
def health():
    try:
        sb = get_supabase()
        sb.table("CLIENTE_INSTALACION").select("POZO_ID").limit(1).execute()
        return jsonify({"ok": True, "message": "Conexión Supabase correcta"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
# ══════════════════════════════════════════════
# IMPORTAR REPORTE INSTALACIÓN
# ══════════════════════════════════════════════

@app.route("/api/importar/requisicion-bienes", methods=["POST"])
def importar_requisicion_bienes():
    registrar_historial_subida(
        sb,
        pozo=pozo_val,
        no_instalacion=no_instalacion,
        tipo="Requisición de Bienes",
        usuario=request.form.get("usuario"),
        archivo=uploaded.filename,
        estado="OK",
        detalle="Importación finalizada"
    )
    try:
        uploaded = request.files.get("file") or request.files.get("archivo")
        if not uploaded:
            return jsonify({"ok": False, "error": "No se envió archivo"}), 400

        sb = get_supabase()

        file_bytes = uploaded.read()
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)

        # Buscar hoja
        if "Requisicion Bienes" in wb.sheetnames:
            ws = wb["Requisicion Bienes"]
        elif len(wb.sheetnames) == 1:
            ws = wb[wb.sheetnames[0]]
        else:
            return jsonify({
                "ok": False,
                "error": "No se encontró la hoja 'Requisicion Bienes'"
            }), 400

        # Leer POZO y FECHA
        pozo_val = None
        fecha_val = None

        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").strip().upper()

                if "POZO/ESTACIÓN" in val or "POZO/ESTACION" in val:
                    pozo_val = _normalize_cell(ws.cell(cell.row, cell.column + 1).value)

                if "FECHA" in val:
                    fecha_val = _to_supabase_value(ws.cell(cell.row, cell.column + 1).value)

            if pozo_val and fecha_val:
                break

        if not pozo_val:
            return jsonify({"ok": False, "error": "No se encontró POZO/ESTACIÓN"}), 400

        no_instalacion = request.form.get("no_instalacion")
        if not no_instalacion:
            return jsonify({"ok": False, "error": "Debe enviar no_instalacion"}), 400

        id_base = f"{pozo_val}-{no_instalacion}"

        # Insertar BIENES_REGISTRO
        registro_payload = {
            "POZO": pozo_val,
            "FECHA": fecha_val,
            "NO_INSTALACION": no_instalacion,
            "ID": id_base,
        }

        ok_reg, err_reg = _safe_write_with_variants(
            sb,
            "BIENES_REGISTRO",
            registro_payload,
            mode="insert"
        )

        if not ok_reg:
            return jsonify({
                "ok": False,
                "error": "Falló BIENES_REGISTRO",
                "detalle": err_reg
            }), 500

        # Leer tabla desde fila 21, columnas C:F
        hdr_row = 21
        start_row = hdr_row + 1

        registros_bienes = 0
        warnings = []

        for r in range(start_row, ws.max_row + 1):
            valores_fila = [
                ws.cell(r, 3).value,
                ws.cell(r, 4).value,
                ws.cell(r, 5).value,
                ws.cell(r, 6).value,
            ]

            texto_fila = " ".join(str(v or "").upper() for v in valores_fila)

            if "OBSERVACIONES" in texto_fila:
                break

            pn_cliente = _normalize_cell(ws.cell(r, 3).value)
            descripcion_cliente = _normalize_cell(ws.cell(r, 4).value)
            pn_syteline = _normalize_cell(ws.cell(r, 5).value)
            descripcion_syteline = _normalize_cell(ws.cell(r, 6).value)

            if not any([pn_cliente, descripcion_cliente, pn_syteline, descripcion_syteline]):
                continue

            bienes_payload = {
                "POZO": pozo_val,
                "PN_CLIENTE": pn_cliente,
                "DESCRIPCION_CLIENTE": descripcion_cliente,
                "PN_SYTELINE": pn_syteline,
                "DESCRIPCION_SYTELINE": descripcion_syteline,
                "NO_INSTALACION": no_instalacion,
                "ID": id_base,
            }

            ok_bien, err_bien = _safe_write_with_variants(
                sb,
                "BIENES",
                bienes_payload,
                mode="insert"
            )

            if ok_bien:
                registros_bienes += 1
            else:
                warnings.append({
                    "fila": r,
                    "error": err_bien,
                    "payload": bienes_payload
                })

        return jsonify({
            "ok": True,
            "message": "Requisición de Bienes importada correctamente",
            "pozo": pozo_val,
            "no_instalacion": no_instalacion,
            "id": id_base,
            "inserted": {
                "BIENES_REGISTRO": 1,
                "BIENES": registros_bienes
            },
            "warnings": warnings
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ══════════════════════════════════════════════
# IMPORTAR REPORTE INSTALACIÓN
# ══════════════════════════════════════════════
@app.route("/api/usuarios", methods=["POST"])
def crear_usuario():
    try:
        data = request.get_json()

        nombre = data.get("nombre")
        apellido = data.get("apellido")
        correo = data.get("correo")
        contrasena = data.get("contrasena")

        if not nombre or not apellido or not correo or not contrasena:
            return jsonify({
                "ok": False,
                "error": "Faltan datos obligatorios"
            }), 400

        sb = get_supabase()

        usuario_payload = {
            "nombre": nombre,
            "apellido": apellido,
            "correo": correo,
            "contrasena": contrasena,
            "permiso_carga": bool(data.get("permiso_carga", False)),
            "permiso_stop": bool(data.get("permiso_stop", False)),
        }
        existe = (
        sb.table("usuarios")
        .select("nombre")
        .eq("nombre", nombre)
        .limit(1)
        .execute()
    )

        if existe.data:
            return jsonify({
                "ok": False,
                "error": "Ya existe un usuario con ese nombre"
            }), 409

        res = sb.table("usuarios").insert(usuario_payload).execute()

        return jsonify({
            "ok": True,
            "message": "Usuario creado correctamente",
            "data": res.data
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500
@app.route("/api/importar/reporte-instalacion", methods=["POST"])
def importar_reporte_instalacion():
       
    registrar_historial_subida(
        sb,
        pozo=pozo_id,
        no_instalacion=instalacion_num,
        tipo="Reporte de Instalación",
        usuario=request.form.get("usuario"),
        archivo=uploaded.filename,
        estado="OK",
        detalle="Importación finalizada"
    )
    try:
        uploaded = request.files.get("file") or request.files.get("archivo")
        if not uploaded:
            return jsonify({"ok": False, "error": "No se envió archivo (use file o archivo)"}), 400

        if not uploaded.filename:
            return jsonify({"ok": False, "error": "Archivo vacío"}), 400

        filename = uploaded.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
            return jsonify({"ok": False, "error": "Solo se permiten archivos .xlsx o .xlsm"}), 400

        sb = get_supabase()

        # Leer Excel desde memoria
        file_bytes = uploaded.read()
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        if "Reporte" not in wb.sheetnames:
            return jsonify({"ok": False, "error": "La hoja 'Reporte' no existe"}), 400

        ws = wb["Reporte"]
        

        warnings = []
        inserted = {}
        debug_steps = []

        # ══════════════════════════════════════════════
        # 1) CLIENTE_INSTALACION
        # ══════════════════════════════════════════════
        vE1  = _cell_value(ws, "E1")   # CLIENTE
        vU1  = _cell_value(ws, "U1")   # CAMPO / BLOQUE
        vAH1 = _cell_value(ws, "AH1")  # POZO (nombre + (id))
        vBB1 = _cell_value(ws, "BB1")  # PAIS

        vE2  = _cell_value(ws, "E2")   # TALADRO
        vQ2  = _cell_value(ws, "Q2")   # FECHA-HORA INICIO
        vZ2  = _cell_value(ws, "Z2")   # FECHA-HORA ARRANQUE
        vAL2 = _cell_value(ws, "AL2")  # FECHA-HORA FINAL
        vAZ2 = _cell_value(ws, "AZ2")  # HORAS TÉCNICO

        vBE2 = _cell_value(ws, "BE2")  # TIPO DE NEGOCIO
        vBO2 = _cell_value(ws, "BO2")  # No. INSTALACIÓN NOVOMET

        vI4  = _cell_value(ws, "I4")   # CONTRATADO POR / SPOOLER

        vY4  = _cell_value(ws, "Y4")
        vY5  = _cell_value(ws, "Y5")
        vAV4 = _cell_value(ws, "AV4")
        vAV5 = _cell_value(ws, "AV5")
        vBM4 = _cell_value(ws, "BM4")
        vBM5 = _cell_value(ws, "BM5")

        campo, bloque = _split_by_slash(vU1)
        pozo, pozo_id = _split_name_parenthesis(vAH1)
        contratado_por, spooler = _split_by_slash(vI4)
        instalacion_text = str(_normalize_cell(vBO2) or "").strip()
        instalacion_num  = _parse_long(instalacion_text)

        if not pozo_id:
            pozo_id = pozo or ""

        id_base = f"{pozo_id}-{instalacion_num}" if instalacion_num is not None else f"{pozo_id}-"

        if not pozo_id:
            _append_warning(warnings, "CLIENTE_INSTALACION", "No se pudo determinar POZO_ID desde AH1", celda="AH1", valor=vAH1)
        if instalacion_num is None:
            _append_warning(warnings, "CLIENTE_INSTALACION", "No se pudo determinar NO_INSTALACION desde BO2", celda="BO2", valor=vBO2)

        cliente_payload = {
            "CLIENTE": _normalize_cell(vE1),
            "CAMPO": campo,
            "BLOQUE": bloque,
            "POZO": _normalize_cell(pozo),
            "POZO_ID": _normalize_cell(pozo_id),
            "NO_INSTALACION": instalacion_num,
            "PAIS": _normalize_cell(vBB1),
            "TALADRO": _normalize_cell(vE2),
            "FECHA_INICIO": _to_supabase_value(vQ2),
            "FECHA_ARRANQUE": _to_supabase_value(vZ2),
            "FECHA_FINAL": _to_supabase_value(vAL2),
            "HORAS_TECNICO": _parse_float(vAZ2),
            "TIPO_NEGOCIO": _normalize_cell(vBE2),
            "CONTRATADO_POR": _normalize_cell(contratado_por),
            "SPOOLER": _normalize_cell(spooler),
            "FECHA_SUBIDA_POLEA_CAPILAR_EXTERNO": _to_supabase_value(vY5),
            "FECHA_SUBIDA_POLEA_CABLE_PROTECTORES": _to_supabase_value(vY4),
            "FECHA_BAJADA_POLEA_CABLE_PROTECTORES": _to_supabase_value(vAV4),
            "FECHA_BAJADA_POLEA_CAPILAR_EXTERNO": _to_supabase_value(vAV5),
            "HORAS_SPOOLER_CABLE_PROTECTORES": _parse_int(vBM4),
            "HORAS_SPOOLER_CAPILAR_EXTERNO": _parse_int(vBM5),
            "ID": id_base,
        }
        # ✅ VALIDAR DUPLICADO ANTES DE INSERTAR
        existe = sb.table("CLIENTE_INSTALACION") \
            .select("ID, POZO_ID, NO_INSTALACION") \
            .eq("ID", id_base) \
            .execute()

        if existe.data:
            return jsonify({
                "ok": False,
                "tipo": "duplicado",
                "mensaje": "No se importó el reporte de instalación porque ya existe en la base de datos.",
                "duplicados": [f"{pozo_id} - Instalación {instalacion_num}"]
            }), 409


        ok_ci, err_ci = _safe_write_with_variants(sb, "CLIENTE_INSTALACION", cliente_payload, mode="insert")        
        inserted["CLIENTE_INSTALACION"] = 1 if ok_ci else 0
        if not ok_ci:
            _append_warning(
                warnings,
                "CLIENTE_INSTALACION",
                "Falló el guardado en Supabase",
                payload=cliente_payload,
                detalle=err_ci,
            )
        else:
            debug_steps.append({"tabla": "CLIENTE_INSTALACION", "estado": "ok", "payload": cliente_payload})

        # ══════════════════════════════════════════════
        # 2) INFPOZO2_INSTALACION
        # Celdas: AB7..AB11 (campos 1-5), AJ7:AY7 (MAX_DLS),
        #         AR8:AU8, AR9:AU9, AR10:AU10, AR11:AY11,
        #         BF7:BO7, BF8:BO8, BF9:BG9, BM9:BO9, BF10:BO11
        # ══════════════════════════════════════════════
        # Campos 1-5: AB7..AB11 (top-left de cada merged block de esa fila)
        vFld = {}
        for rr in range(7, 12):
            vFld[rr - 6] = _cell_value(ws, f"AB{rr}")

        # Campo 6: concatenar AJ7:AY7 respetando bloques merged
        col_AJ = column_index_from_string("AJ")
        col_AY = column_index_from_string("AY")
        vMAX_DLS = _concat_merged_row(ws, 7, col_AJ, col_AY, sep=" ")
        if vMAX_DLS is None:
            vMAX_DLS = _normalize_cell(_cell_value(ws, "AJ7"))

        # Campos 7-10: AR8:AU8, AR9:AU9, AR10:AU10, AR11:AY11
        # (top-left de cada rango -> AR8, AR9, AR10, AR11)
        vDLS_PROF_BOMBA  = _normalize_cell(_cell_value(ws, "AR8"))
        vDESV_PROF_BOMBA = _normalize_cell(_cell_value(ws, "AR9"))
        vMAX_DESV_RUN    = _normalize_cell(_cell_value(ws, "AR10"))
        vPROF_INTAKE     = _normalize_cell(_cell_value(ws, "AR11"))

        # Campos 11-15: BF7:BO7, BF8:BO8, BF9:BG9, BM9:BO9, BF10:BO11
        vTOPE_PERF_MD       = _normalize_cell(_cell_value(ws, "BF7"))
        vZONA_PROD_INI      = _normalize_cell(_cell_value(ws, "BF8"))
        vNo_WORKOVER        = _normalize_cell(_cell_value(ws, "BF9"))
        vLOG_EQUIPO         = _normalize_cell(_cell_value(ws, "BM9"))
        vPESO_SARTA_SUBIENDO = _normalize_cell(_cell_value(ws, "BF10"))

        vAPI = _parse_float(vFld.get(1))
        vCORTE_AGUA    = _normalize_cell(vFld.get(2))
        vGOR = _to_int(vFld.get(3))
        vTEMP_CABEZA = _parse_int(vFld.get(4))
        vTEMP_RESERV = _parse_int(vFld.get(5))

        infpozo2_payload = {
            "ID": id_base,
            "POZO_ID": _normalize_cell(pozo_id),
            "NO_INSTALACION": instalacion_num,
            "API": vAPI,
            "CORTE_AGUA": vCORTE_AGUA,
            "GOR": vGOR,
            "TEMP_CABEZA": _parse_int(vFld.get(4)),
            "TEMP_RESERVORIO": _parse_int(vFld.get(5)),
            "MAX_DLS_PROFUNDIDAD": _normalize_cell(vMAX_DLS),
            "DLS_PROF_BOMBA": vDLS_PROF_BOMBA,
            "DESVIACION_PROF_BOMBA": vDESV_PROF_BOMBA,
            "MAX_DESVIACION_RUNNING": vMAX_DESV_RUN,
            "PROFUNDIDAD_INTAKE_MD": vPROF_INTAKE,
            "TOPE_PERFORACION_MD": vTOPE_PERF_MD,
            "ZONA_PRODUCTORA_INICIAL": vZONA_PROD_INI,
            "NO_WORKOVER": vNo_WORKOVER,
            "LOG_EQUIPO": _parse_int(vLOG_EQUIPO),
            "PESO_SARTA_SUBIENDO": vPESO_SARTA_SUBIENDO,
        }
        ok_ip2, err_ip2 = _safe_write_with_variants(sb, "INFPOZO2_INSTALACION", infpozo2_payload, mode="upsert", on_conflict="ID")
        inserted["INFPOZO2_INSTALACION"] = 1 if ok_ip2 else 0
        if not ok_ip2:
            _append_warning(
                warnings,
                "INFPOZO2_INSTALACION",
                "Falló el guardado en Supabase",
                payload=infpozo2_payload,
                detalle=err_ip2,
            )
        else:
            debug_steps.append({"tabla": "INFPOZO2_INSTALACION", "estado": "ok", "payload": infpozo2_payload})

        # ══════════════════════════════════════════════
        # 3) INFPOZO_INSTALACION  (filas 8..11, columnas A..W)
        # ══════════════════════════════════════════════
        start_col, end_col = ws["A8"].column, ws["W8"].column
        logical_starts = _extract_textbook_rows(ws, 8, 11, start_col, end_col)

        def _read_block(row, col):
            return _normalize_cell(_merge_top_left_text(ws, row, col))

        for r in range(8, 12):
            row_values = [_read_block(r, logical_starts[i]) for i in range(6)]
            if all(v in (None, "") for v in row_values):
                debug_steps.append({"tabla": "INFPOZO_INSTALACION", "fila": r, "estado": "omitido", "motivo": "fila vacía"})
                continue

            infpozo_payload = {
                "ID": id_base,
                "POZO_ID": _normalize_cell(pozo_id),
                "NO_INSTALACION": instalacion_num,
                "EQUIPO": row_values[0],
                "TOP_MD_FT": row_values[1],
                "OD_IN": row_values[2],
                "PESO_LB_FT": row_values[3],
                "CANTIDAD_JTS": row_values[4],
                "CLASE_A_B": row_values[5],
            }
            ok_ip, err_ip = _safe_insert(sb, "INFPOZO_INSTALACION", infpozo_payload)
            inserted["INFPOZO_INSTALACION"] = inserted.get("INFPOZO_INSTALACION", 0) + (1 if ok_ip else 0)
            if not ok_ip:
                warnings.append({"tabla": "INFPOZO_INSTALACION", "fila": r, "error": err_ip})

        # ══════════════════════════════════════════════
        # 4) EQUIPOFONDO_INSTALACION
        # Encabezado en fila 13; datos hasta primera fila vacía.
        # La fila siguiente al último dato + 1 = encabezado de EQUISUPERFICIE.
        # ══════════════════════════════════════════════
        last_row   = ws.max_row
        hdr_e_row  = 13
        max_col_e  = ws["BO1"].column
        hdr_cols_e = _detect_header_cols(ws, hdr_e_row, max_col_e)

        hdr_es_row        = None
        last_inserted_e_row = 0

        for r in range(hdr_e_row + 1, last_row + 1):
            row_vals = [
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["EQUIPO"]))   if hdr_cols_e["EQUIPO"]   else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["PARTE"]))    if hdr_cols_e["PARTE"]    else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["SERIE"]))    if hdr_cols_e["SERIE"]    else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["REEL"]))     if hdr_cols_e["REEL"]     else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["DESCRIP"]))  if hdr_cols_e["DESCRIP"]  else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["CONDIC"]))   if hdr_cols_e["CONDIC"]   else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["PROPIED"]))  if hdr_cols_e["PROPIED"]  else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["OD"]))       if hdr_cols_e["OD"]       else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["LONGITUD"])) if hdr_cols_e["LONGITUD"] else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["PROF"]))     if hdr_cols_e["PROF"]     else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["SHIMS"]))    if hdr_cols_e["SHIMS"]    else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_e["GIRO"]))     if hdr_cols_e["GIRO"]     else None,
            ]

            if all(v in (None, "") for v in row_vals):
                # La fila vacía marca el fin del bloque; el header de EQUISUPERFICIE
                # está en la siguiente fila (igual que VBA: hdrESRow = rE + 1)
                hdr_es_row = r + 1
                break

            payload = {
                "ID": id_base,
                "POZO_ID": _normalize_cell(pozo_id),
                "NO_INSTALACION": instalacion_num,
                "EQUIPO": row_vals[0],
                "NO_PARTE": row_vals[1],
                "NO_SERIE": row_vals[2],
                "NO_REEL": row_vals[3],
                "DESCRIPCION": row_vals[4],
                "CONDICION": row_vals[5],
                "PROPIEDAD": row_vals[6],
                "OD": row_vals[7],
                "LONGITUD": row_vals[8],
                "PROF_TOP_MD": row_vals[9],
                "SHIMS_BASE": row_vals[10],
                "GIRO": row_vals[11],
            }
            ok, err = _safe_insert(sb, "EQUIPOFONDO_INSTALACION", payload)
            inserted["EQUIPOFONDO_INSTALACION"] = inserted.get("EQUIPOFONDO_INSTALACION", 0) + (1 if ok else 0)
            if not ok:
                warnings.append({"tabla": "EQUIPOFONDO_INSTALACION", "fila": r, "error": err})
            else:
                last_inserted_e_row = r

        # Calcular hdr_es_row igual que VBA:
        # Si hubo registros insertados: hdrESRow = lastInsertedERow + 2
        # Si no: usar el valor detectado por la fila vacía, o fallback
        if last_inserted_e_row > 0:
            hdr_es_row = last_inserted_e_row + 2
        if hdr_es_row is None:
            hdr_es_row = last_row + 1

        # ══════════════════════════════════════════════
        # 5) EQUISUPERFICIE_INSTALACION
        # ══════════════════════════════════════════════
        if hdr_es_row is None or hdr_es_row <= 0:
            _append_warning(
                warnings,
                "EQUISUPERFICIE_INSTALACION",
                "No se pudo detectar la fila de encabezado; se usará un valor de respaldo",
                hdr_es_row=hdr_es_row,
                last_inserted_e_row=last_inserted_e_row,
            )
            hdr_es_row = max(1, hdr_acc_row + 1)

        hdr_es_row     = min(max(hdr_es_row, 1), last_row)
        hdr_cols_es    = _detect_header_cols(ws, hdr_es_row, ws.max_column)

        acc_hdr_row             = None
        last_inserted_es_row    = 0

        for r in range(hdr_es_row + 1, last_row + 1):
            # Detener si aparece COMENTARIOS en la fila (marca inicio de ACCESORIOS/OBS)
            found_comentarios = any(
                isinstance(_normalize_cell(_merge_top_left_text(ws, r, c)), str)
                and "COMENTARIOS" in (_normalize_cell(_merge_top_left_text(ws, r, c)) or "").upper()
                for c in range(1, min(ws.max_column, 5) + 1)
            )
            if found_comentarios:
                acc_hdr_row = r
                break

            row_vals = [
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_es["EQUIPO"]))  if hdr_cols_es["EQUIPO"]  else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_es["PARTE"]))   if hdr_cols_es["PARTE"]   else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_es["SERIE"]))   if hdr_cols_es["SERIE"]   else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_es["DESCRIP"])) if hdr_cols_es["DESCRIP"] else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_es["CONDIC"]))  if hdr_cols_es["CONDIC"]  else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_es["PROPIED"])) if hdr_cols_es["PROPIED"] else None,
            ]

            if all(v in (None, "") for v in row_vals):
                acc_hdr_row = r + 1
                break

            payload = {
                "ID": id_base,
                "POZO_ID": _normalize_cell(pozo_id),
                "NO_INSTALACION": instalacion_num,
                "EQUIPO": row_vals[0],
                "NO_PARTE": row_vals[1],
                "NO_SERIE": row_vals[2],
                "DESCRIPCION": row_vals[3],
                "CONDICION": row_vals[4],
                "PROPIEDAD": row_vals[5],
            }
            ok, err = _safe_insert(sb, "EQUISUPERFICIE_INSTALACION", payload)
            inserted["EQUISUPERFICIE_INSTALACION"] = inserted.get("EQUISUPERFICIE_INSTALACION", 0) + (1 if ok else 0)
            if not ok:
                warnings.append({"tabla": "EQUISUPERFICIE_INSTALACION", "fila": r, "error": err})
            else:
                last_inserted_es_row = r

        # Recalcular acc_hdr_row igual que VBA
        if last_inserted_es_row > 0:
            acc_hdr_row = last_inserted_es_row + 2
        if acc_hdr_row is None:
            acc_hdr_row = hdr_es_row + 1

        # ══════════════════════════════════════════════
        # 6) ACCESORIOS_INSTALACION
        # ══════════════════════════════════════════════
        hdr_acc_row   = acc_hdr_row
        hdr_cols_acc  = _detect_header_cols(ws, hdr_acc_row, ws.max_column)
        last_inserted_acc_row = 0

        for r in range(hdr_acc_row + 1, last_row + 1):
            # Detener si aparece la palabra COMENTARIOS en cualquier columna de la fila
            found_comentarios = False
            for c in range(1, ws.max_column + 1):
                txt = _normalize_cell(_merge_top_left_text(ws, r, c))
                if isinstance(txt, str) and "COMENTARIOS" in txt.upper():
                    found_comentarios = True
                    break
            if found_comentarios:
                break

            row_vals = [
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_acc["ACCESORIOS"])) if hdr_cols_acc["ACCESORIOS"] else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_acc["PARTE"]))      if hdr_cols_acc["PARTE"]      else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_acc["DESCRIP"]))    if hdr_cols_acc["DESCRIP"]    else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_acc["CANTIDAD"]))   if hdr_cols_acc["CANTIDAD"]   else None,
                _normalize_cell(_merge_top_left_text(ws, r, hdr_cols_acc["PROPIED"]))    if hdr_cols_acc["PROPIED"]    else None,
            ]
            if all(v in (None, "") for v in row_vals):
                break

            payload = {
                "ID": id_base,
                "POZO_ID": _normalize_cell(pozo_id),
                "NO_INSTALACION": instalacion_num,
                "ACCESORIOS": row_vals[0],
                "NO_PARTE": row_vals[1],
                "DESCRIPCION": row_vals[2],
                "CANTIDAD": _parse_long(row_vals[3]),
                "PROPIEDAD": row_vals[4],
            }
            ok, err = _safe_insert(sb, "ACCESORIOS_INSTALACION", payload)
            inserted["ACCESORIOS_INSTALACION"] = inserted.get("ACCESORIOS_INSTALACION", 0) + (1 if ok else 0)
            if not ok:
                warnings.append({"tabla": "ACCESORIOS_INSTALACION", "fila": r, "error": err})
            else:
                last_inserted_acc_row = r

        # ══════════════════════════════════════════════
        # 7) PARAMVSD_INSTALACION
        # El VBA lee PARAMETROS VSD/SUT en los mismos grupos de filas del bloque
        # de EQUIPO_SUPERFICIE (columnas derechas AV:BN), empezando en hdrESRow + 2.
        # Grupo 1: AV:AW → PARAMETRO | AX:BA → VALOR | BB:BE → ESTADO  (8 filas)
        # Grupo 2: BF:BG → PARAMETRO | BH:BJ → VALOR | BK:BN → ESTADO  (8 filas)
        #   Excepción fila 41 en Grupo 2: BH:BN → VALOR, ESTADO = ""
        # ══════════════════════════════════════════════
        try:
            fila_vsd_inicio = (hdr_es_row + 2) if hdr_es_row > 0 else 39

            # Grupo 1
            for fila in range(fila_vsd_inicio, fila_vsd_inicio + 8):
                parametro = _concat_range_raw(ws, fila, "AV", "AW")
                valor     = _concat_range_raw(ws, fila, "AX", "BA")
                estado    = _concat_range_raw(ws, fila, "BB", "BE")

                payload = {
                    "ID": id_base,
                    "POZO_ID": _normalize_cell(pozo_id),
                    "NO_INSTALACION": instalacion_num,
                    "PARAMETRO": parametro or None,
                    "VALOR": valor or None,
                    "ESTADO": estado or None,
                }
                ok, err = _safe_insert(sb, "PARAMVSD_INSTALACION", payload)
                inserted["PARAMVSD_INSTALACION"] = inserted.get("PARAMVSD_INSTALACION", 0) + (1 if ok else 0)
                if not ok:
                    warnings.append({"tabla": "PARAMVSD_INSTALACION", "grupo": 1, "fila": fila, "error": err})

            # Grupo 2
            for fila in range(fila_vsd_inicio, fila_vsd_inicio + 8):
                parametro = _concat_range_raw(ws, fila, "BF", "BG")

                if fila == 41:
                    # Regla especial: BH:BN → VALOR; ESTADO vacío
                    valor  = _concat_range_raw(ws, fila, "BH", "BN")
                    estado = ""
                else:
                    valor  = _concat_range_raw(ws, fila, "BH", "BJ")
                    estado = _concat_range_raw(ws, fila, "BK", "BN")

                payload = {
                    "ID": id_base,
                    "POZO_ID": _normalize_cell(pozo_id),
                    "NO_INSTALACION": instalacion_num,
                    "PARAMETRO": parametro or None,
                    "VALOR": valor or None,
                    "ESTADO": estado or None,
                }
                ok, err = _safe_insert(sb, "PARAMVSD_INSTALACION", payload)
                inserted["PARAMVSD_INSTALACION"] = inserted.get("PARAMVSD_INSTALACION", 0) + (1 if ok else 0)
                if not ok:
                    warnings.append({"tabla": "PARAMVSD_INSTALACION", "grupo": 2, "fila": fila, "error": err})

        except Exception as e_vsd:
            warnings.append({"tabla": "PARAMVSD_INSTALACION", "error": str(e_vsd)})

        # ══════════════════════════════════════════════
        # 8) PARAMESTATIC_INSTALACION
        # Empieza en accHeaderRow + 4 (4 filas después del header de ACCESORIOS).
        # Columnas (respetando merged areas):
        #   AB:AG → PRUEBA
        #   AH    → FECHA (parte date)   AK → HORA (parte time)  → combinadas → FECHA
        #   AM:AO → PIP_psi
        #   AP:AR → PD_psi
        #   AS:AU → T_INTAKE
        #   AV:AX → T_MOTOR
        #   AY:BB → F_F
        #   BC:BD → F_T
        #   BE:BF → C_L
        #   BG:BJ → AMP_MOTOR
        #   BK:BL → HZ
        #   BM:BO → BSW_BFPH
        # Avance de fila: basado en el MergeArea de la columna AB de la fila actual.
        # ══════════════════════════════════════════════
        try:
            col_AB = column_index_from_string("AB")
            col_AH = column_index_from_string("AH")
            col_AK = column_index_from_string("AK")

            if not acc_hdr_row:
                _append_warning(
                    warnings,
                    "PARAMESTATIC_INSTALACION",
                    "No se detectó la fila de encabezado de ACCESORIOS; se usará un respaldo",
                    acc_hdr_row=acc_hdr_row,
                    hdr_es_row=hdr_es_row,
                )
                acc_hdr_row = hdr_es_row + 2

            fila_a    = acc_hdr_row + 4
            any_arr   = 0
            max_arr   = 10   # máximo de filas de parámetros a leer

            while any_arr < max_arr:
                if fila_a > last_row:
                    break

                # Leer PRUEBA
                v_prueba = _concat_range_raw(ws, fila_a, "AB", "AG")
                v_prueba = _normalize_cell(v_prueba)

                # Leer FECHA (AH) y HORA (AK) por separado y combinar
                raw_fecha = _merge_top_left_text(ws, fila_a, col_AH)
                raw_hora  = _merge_top_left_text(ws, fila_a, col_AK)
                v_fecha   = _combine_excel_date_time(raw_fecha, raw_hora)

                # Leer resto de columnas
                v_pip      = _normalize_cell(_concat_range_raw(ws, fila_a, "AM", "AO"))
                v_pd       = _normalize_cell(_concat_range_raw(ws, fila_a, "AP", "AR"))
                v_t_intake = _normalize_cell(_concat_range_raw(ws, fila_a, "AS", "AU"))
                v_t_motor  = _normalize_cell(_concat_range_raw(ws, fila_a, "AV", "AX"))
                v_f_f      = _normalize_cell(_concat_range_raw(ws, fila_a, "AY", "BB"))
                v_f_t      = _normalize_cell(_concat_range_raw(ws, fila_a, "BC", "BD"))
                v_c_l      = _normalize_cell(_concat_range_raw(ws, fila_a, "BE", "BF"))
                v_amp      = _normalize_cell(_concat_range_raw(ws, fila_a, "BG", "BJ"))
                v_hz       = _normalize_cell(_concat_range_raw(ws, fila_a, "BK", "BL"))
                v_bsw      = _normalize_cell(_concat_range_raw(ws, fila_a, "BM", "BO"))

                # Si todos los campos están vacíos, saltar fila (no insertar)
                all_empty = all(v in (None, "") for v in [
                    v_prueba, v_fecha, v_pip, v_pd, v_t_intake, v_t_motor,
                    v_f_f, v_f_t, v_c_l, v_amp, v_hz, v_bsw
                ])

                if not all_empty:
                    payload = {
                        
                        "POZO_ID": _normalize_cell(pozo_id),
                        "NO_INSTALACION": instalacion_num,
                        "PRUEBA": v_prueba,
                        "FECHA": v_fecha,
                        "PIP_PSI": v_pip,
                        "PD_PSI": v_pd,
                        "T_INTAKE": v_t_intake,
                        "T_MOTOR": v_t_motor,
                        "F_F": v_f_f,
                        "F_T": v_f_t,
                        "C_L": v_c_l,
                        "AMP_MOTOR": v_amp,
                        "HZ": v_hz,
                        "BSW_BFPH": v_bsw,
                        "ID": id_base,
                    }
                    ok, err = _safe_write_with_variants(sb, "PARAMESTATIC_INSTALACION", payload, mode="insert")
                    inserted["PARAMESTATIC_INSTALACION"] = inserted.get("PARAMESTATIC_INSTALACION", 0) + (1 if ok else 0)
                    if not ok:
                        _append_warning(
                            warnings,
                            "PARAMESTATIC_INSTALACION",
                            "Falló el guardado en Supabase",
                            fila=fila_a,
                            payload=payload,
                            detalle=err,
                        )
                    else:
                        if isinstance(err, dict) and err.get("dropped_due_error"):
                            _append_warning(
                                warnings,
                                "PARAMESTATIC_INSTALACION",
                                "Se omitieron columnas que no existen en la tabla real durante el guardado",
                                fila=fila_a,
                                columnas_omitidas=err.get("dropped_due_error"),
                            )
                        any_arr += 1
                        debug_steps.append({"tabla": "PARAMESTATIC_INSTALACION", "fila": fila_a, "estado": "ok", "payload": payload})
                else:
                    debug_steps.append({"tabla": "PARAMESTATIC_INSTALACION", "fila": fila_a, "estado": "omitido", "motivo": "todos los campos vacíos"})

                # Avance de fila: respetar MergeArea de columna AB (igual que VBA)
                cell_ab = ws.cell(row=fila_a, column=col_AB)
                merged_ab = None
                for rng in ws.merged_cells.ranges:
                    if cell_ab.coordinate in rng:
                        merged_ab = rng
                        break
                if merged_ab is not None:
                    fila_a = merged_ab.max_row + 1
                else:
                    fila_a += 1

        except Exception as e_arr:
            warnings.append({"tabla": "PARAMESTATIC_INSTALACION", "error": str(e_arr)})

        # ══════════════════════════════════════════════
        # 9) OBSERVACIONES_INSTALACION
        # Localiza la sección "COMENTARIOS" después del último registro de ACCESORIOS.
        # Extrae: COMENTARIOS, MATERIAL_UTILIZADO, MATERIAL_SOBRANTE, NOTAS, TECNICO_NOVOMET.
        # TECNICO: escanea de fila 200 hacia abajo en cols AV:BO buscando última fila con datos.
        # ══════════════════════════════════════════════
        try:
            # ── Localizar encabezado de COMENTARIOS ───────────────────────
            obs_hdr_row      = 0
            obs_col_comentarios = 0

            # Primero intenta desde la fila siguiente al último ACC insertado
            search_start = (last_inserted_acc_row + 1) if last_inserted_acc_row > 0 else hdr_acc_row

            for obs_r in range(search_start, last_row + 1):
                for obs_c in range(1, min(ws.max_column, ws.max_column) + 1):
                    val = _merge_top_left_text(ws, obs_r, obs_c)
                    if val is None:
                        continue
                    val_upper = str(val).strip().upper()
                    if "COMENTARIOS" in val_upper:
                        obs_hdr_row      = obs_r
                        obs_col_comentarios = obs_c
                        break
                if obs_hdr_row > 0:
                    break

            if obs_hdr_row == 0:
                warnings.append({
                    "tabla": "OBSERVACIONES_INSTALACION",
                    "error": "No se detectó la sección COMENTARIOS; se omitió la importación."
                })
            else:
                # ── Detectar sub-secciones ─────────────────────────────────
                obs_row_mat_util  = 0
                obs_row_mat_sobr  = 0
                obs_row_notas     = 0
                obs_col_mat_util  = obs_col_comentarios
                obs_col_mat_sobr  = obs_col_comentarios
                obs_col_notas     = obs_col_comentarios

                for obs_r in range(obs_hdr_row, last_row + 1):
                    for obs_c in range(1, ws.max_column + 1):
                        val = _merge_top_left_text(ws, obs_r, obs_c)
                        if val is None:
                            continue
                        val_upper = str(val).strip().upper()
                        if not val_upper:
                            continue
                        if obs_row_mat_util == 0 and "MATERIAL UTILIZADO" in val_upper:
                            obs_row_mat_util = obs_r
                            obs_col_mat_util = obs_c
                        if obs_row_mat_sobr == 0 and ("MATERIAL SOBRANTE" in val_upper or "MATERIAL SOBRANTES" in val_upper):
                            obs_row_mat_sobr = obs_r
                            obs_col_mat_sobr = obs_c
                        if obs_row_notas == 0 and val_upper.startswith("NOTA"):
                            obs_row_notas = obs_r
                            obs_col_notas = obs_c

                obs_data_start = obs_hdr_row + 1

                # Fallbacks si no se encontraron sub-secciones
                if obs_row_mat_util == 0:
                    obs_row_mat_util = obs_data_start
                if obs_row_mat_sobr == 0:
                    obs_row_mat_sobr = obs_row_mat_util + 1
                if obs_row_notas == 0:
                    obs_row_notas = obs_row_mat_sobr + 1

                # ── Extraer COMENTARIOS (obs_data_start .. obs_row_mat_util - 1) ──
                comentarios_lines = []
                for obs_i in range(obs_data_start, obs_row_mat_util):
                    val = _normalize_cell(_merge_top_left_text(ws, obs_i, obs_col_comentarios))
                    if val and str(val).strip():
                        comentarios_lines.append(str(val).strip())
                obs_comentarios = "\n".join(comentarios_lines) if comentarios_lines else None

                # ── Extraer MATERIAL_UTILIZADO ─────────────────────────────
                mat_util_lines = []
                for obs_i in range(obs_row_mat_util + 1, obs_row_mat_sobr):
                    val = _normalize_cell(_merge_top_left_text(ws, obs_i, obs_col_mat_util))
                    if val and str(val).strip():
                        mat_util_lines.append(str(val).strip())
                obs_mat_util = "\n".join(mat_util_lines) if mat_util_lines else None

                # ── Extraer MATERIAL_SOBRANTE ──────────────────────────────
                mat_sobr_lines = []
                for obs_i in range(obs_row_mat_sobr + 1, obs_row_notas):
                    val = _normalize_cell(_merge_top_left_text(ws, obs_i, obs_col_mat_sobr))
                    if val and str(val).strip():
                        mat_sobr_lines.append(str(val).strip())
                obs_mat_sobr = "\n".join(mat_sobr_lines) if mat_sobr_lines else None

                # ── Extraer NOTAS (hasta primera fila vacía) ───────────────
                notas_lines = []
                for obs_i in range(obs_row_notas + 1, last_row + 1):
                    # Verificar si la fila tiene datos en alguna columna
                    row_has_data = any(
                        _normalize_cell(_merge_top_left_text(ws, obs_i, obs_c)) not in (None, "")
                        for obs_c in range(1, ws.max_column + 1)
                    )
                    if not row_has_data:
                        break
                    val = _normalize_cell(_merge_top_left_text(ws, obs_i, obs_col_notas))
                    if val and str(val).strip():
                        notas_lines.append(str(val).strip())
                obs_notas = "\n".join(notas_lines) if notas_lines else None

                # ── Extraer TECNICO_NOVOMET ────────────────────────────────
                # Escanea desde fila 200 hacia abajo buscando la última fila con datos
                # en columnas AV:BO (igual que el VBA con OBS_SEARCH_LIMIT = 200).
                col_AV_obs = column_index_from_string("AV")
                col_BO_obs = column_index_from_string("BO")
                obs_last_data_row = 0

                for obs_i in range(200, 0, -1):
                    row_has_data = False
                    for obs_c in range(col_AV_obs, col_BO_obs + 1):
                        val = _normalize_cell(ws.cell(row=obs_i, column=obs_c).value)
                        if val not in (None, ""):
                            val_upper = str(val).upper()
                            if "NOTA:" not in val_upper:
                                row_has_data = True
                                break
                    if row_has_data:
                        obs_last_data_row = obs_i
                        break

                obs_tecnico = None
                if obs_last_data_row > 0:
                    # Leer el valor en la última fila con datos (cols AV:BO)
                    for obs_c in range(col_AV_obs, col_BO_obs + 1):
                        val = _normalize_cell(_merge_top_left_text(ws, obs_last_data_row, obs_c))
                        if val and str(val).strip():
                            val_upper = str(val).upper()
                            if "NOTA:" not in val_upper and "TECNICO NOVOMET" not in val_upper and "TÉCNICO NOVOMET" not in val_upper:
                                obs_tecnico = str(val).strip()
                                break

                # ── Insertar OBSERVACIONES_INSTALACION ────────────────────
                obs_payload = {
                    "ID": id_base,
                    "POZO_ID": _normalize_cell(pozo_id),
                    "NO_INSTALACION": instalacion_num,
                    "COMENTARIOS": obs_comentarios,
                    "MATERIAL_UTILIZADO": obs_mat_util,
                    "MATERIAL_SOBRANTE": obs_mat_sobr,
                    "NOTAS": obs_notas,
                    "TECNICO_NOVOMET": obs_tecnico,
                }
                ok_obs, err_obs = _safe_insert(sb, "OBSERVACIONES_INSTALACION", obs_payload)
                inserted["OBSERVACIONES_INSTALACION"] = 1 if ok_obs else 0
                if not ok_obs:
                    warnings.append({"tabla": "OBSERVACIONES_INSTALACION", "error": err_obs})

        except Exception as e_obs:
            warnings.append({"tabla": "OBSERVACIONES_INSTALACION", "error": str(e_obs)})

        # ══════════════════════════════════════════════
        # Respuesta final
        # ══════════════════════════════════════════════
        total_inserted = sum(inserted.values())
        status_ok      = len(warnings) == 0 and total_inserted > 0

        return jsonify({
            "ok": status_ok,
            "message": "Importación finalizada" if status_ok else "Importación finalizada con advertencias",
            "pozo_id": pozo_id,
            "no_instalacion": instalacion_num,
            "id": id_base,
            "inserted": inserted,
            "total_inserted": total_inserted,
            "warnings": warnings,
            "debug_steps": debug_steps,
        }), (200 if total_inserted > 0 else 500)

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# CLIENTE_INSTALACION
# ══════════════════════════════════════════════

@app.route("/api/cliente_instalacion", methods=["GET"])
def cliente_instalacion():
    try:
        sb = get_supabase()
        res = (
            sb.table("CLIENTE_INSTALACION")
            .select("POZO_ID,NO_INSTALACION,CLIENTE,BLOQUE,CAMPO")
            .not_.is_("POZO_ID", "null")
            .order("POZO_ID")
            .execute()
        )
        return jsonify({"ok": True, "data": serialize_rows(res.data)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# STATUS
# ══════════════════════════════════════════════
@app.route("/api/status", methods=["GET"])
def obtener_status():
    try:
        sb = get_supabase()

        res = (
            sb.table("STATUS")
            .select("POZO_ID,NO_INSTALACION,STOP_DATE,RAZON_STOP,START_DATE")
            .not_.is_("STOP_DATE", "null")
            .is_("START_DATE", "null")  # 👈 solo abiertos
            .order("STOP_DATE", desc=True)
            .execute()
        )

        return jsonify({"ok": True, "data": serialize_rows(res.data)})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# CLIENTE_PULL
# ══════════════════════════════════════════════

@app.route("/api/cliente_pull", methods=["GET"])
def cliente_pull():
    try:
        sb = get_supabase()
        res = (
            sb.table("CLIENTE_PULL")
            .select("POZO_ID,FECHA_INICIO,FECHA_PARADA,NUM_PULL,RAZON_PULL")
            .not_.is_("POZO_ID", "null")
            .order("FECHA_INICIO", desc=True)
            .execute()
        )
        return jsonify({"ok": True, "data": serialize_rows(res.data)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# IMPORTAR REPORTE PULLING
# ══════════════════════════════════════════════
@app.route("/api/importar/reporte-pulling", methods=["POST"])
def importar_reporte_pulling():
    registrar_historial_subida(
        sb,
        pozo=pozo_id,
        no_instalacion=selected_no_inst_payload,
        tipo="Reporte de Pulling",
        usuario=request.form.get("usuario"),
        archivo=uploaded.filename,
        estado="OK",
        detalle="Importación finalizada"
    )

    try:
        uploaded = request.files.get("file") or request.files.get("archivo")
        if not uploaded:
            return jsonify({"ok": False, "error": "No se envió archivo (use file o archivo)"}), 400

        if not uploaded.filename:
            return jsonify({"ok": False, "error": "Archivo vacío"}), 400

        filename = uploaded.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
            return jsonify({"ok": False, "error": "Solo se permiten archivos .xlsx o .xlsm"}), 400

        no_instalacion_raw = (
            request.form.get("no_instalacion")
            or request.args.get("no_instalacion")
            or (request.get_json(silent=True) or {}).get("no_instalacion")
        )
        if no_instalacion_raw is None or str(no_instalacion_raw).strip() == "":
            return jsonify({
                "ok": False,
                "error": "Falta NO_INSTALACION",
                "mensaje": "Agrega el valor de NO_INSTALACION antes de importar el archivo. Puedes enviarlo como form field, query param o JSON.",
                "campo_requerido": "no_instalacion",
                "ejemplo_form": {"no_instalacion": "12345"}
            }), 400

        selected_no_inst = str(no_instalacion_raw).strip()
        selected_no_inst_num = _parse_long(selected_no_inst)
        selected_no_inst_payload = selected_no_inst_num if selected_no_inst_num is not None else selected_no_inst

        sb = get_supabase()

        file_bytes = uploaded.read()
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        if "PULLING" not in wb.sheetnames:
            return jsonify({"ok": False, "error": "La hoja 'PULLING' no existe"}), 400

        ws = wb["PULLING"]

        warnings = []
        inserted = {}
        debug_steps = []

        def _merged_range_for_cell(sheet, row: int, col: int):
            coord = sheet.cell(row=row, column=col).coordinate
            for rng in sheet.merged_cells.ranges:
                if coord in rng:
                    return rng
            return None

        def _merged_top_row(sheet, row: int, col: int) -> int:
            rng = _merged_range_for_cell(sheet, row, col)
            return rng.min_row if rng is not None else row

        def _merged_top_left_value(sheet, row: int, col: int):
            return _cell_value(sheet, sheet.cell(row=row, column=col).coordinate)

        def _clean_header(value) -> str:
            return _normalize_header(value or "")

        def _date_only_value(value):
            value = _normalize_cell(value)
            if value is None:
                return None
            if isinstance(value, datetime):
                return value.date().strftime("%Y-%m-%d")
            if isinstance(value, date):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, str):
                s = value.strip()
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        return datetime.strptime(s, fmt).date().strftime("%Y-%m-%d")
                    except Exception:
                        pass
                return s[:10] if len(s) >= 10 else s
            return value

        # ══════════════════════════════════════════════
        # 1) CLIENTE_PULL
        # ══════════════════════════════════════════════
        vU1 = _cell_value(ws, "R1")
        vAH1 = _cell_value(ws, "J1")
        vCliente = _cell_value(ws, "C1")
        vRig = _cell_value(ws, "AF1")
        vNegocio = _cell_value(ws, "AQ1")
        vNumPull = _cell_value(ws, "AI3")
        vTiempoEqui = _cell_value(ws, "AI2")
        vTecnico1 = _cell_value(ws, "AF112")
        vTecnico2 = _cell_value(ws, "AM112")
        vMotivoParada = _cell_value(ws, "AB4")
        vRazonPull = _cell_value(ws, "AB5")
        vCiudad = _cell_value(ws, "Z1")

        vL2 = _cell_value(ws, "L2")
        vL3 = _cell_value(ws, "L3")
        vL4 = _cell_value(ws, "L4")
        vL5 = _cell_value(ws, "L5")

        vQ2 = _cell_value(ws, "Q2")
        vQ4 = _cell_value(ws, "Q4")
        vQ5 = _cell_value(ws, "Q5")

        vSpooler = _cell_value(ws, "G8")
        vFechaCable = _cell_value(ws, "P8")
        vFechaCap = _cell_value(ws, "P9")
        vFechaCableEnd = _cell_value(ws, "AC8")
        vFechaCapEnd = _cell_value(ws, "AC9")
        vHorasCable = _cell_value(ws, "AQ8")
        vHorasCap = _cell_value(ws, "AQ9")

        campo, bloque = _split_by_slash(vU1)
        pozo, pozo_id = _split_name_parenthesis(vAH1)
        if not pozo_id:
            pozo_id = pozo or ""

        tecnicos = _normalize_cell(vTecnico1) or ""
        tecnico2 = _normalize_cell(vTecnico2)
        if tecnico2 not in (None, ""):
            if tecnicos:
                tecnicos = f"{tecnicos} / "
            tecnicos = f"{tecnicos}{tecnico2}"

        id_base = f"{pozo_id}-{selected_no_inst}" if pozo_id and selected_no_inst else f"{pozo_id}-"
        if not pozo_id:
            _append_warning(warnings, "CLIENTE_PULL", "No se pudo determinar POZO_ID desde J1", celda="J1", valor=vAH1)
        if not selected_no_inst:
            _append_warning(warnings, "CLIENTE_PULL", "NO_INSTALACION vacío", valor=no_instalacion_raw)

        cliente_payload = {
            "ID": id_base,
            "CLIENTE": _normalize_cell(vCliente),
            "POZO": _normalize_cell(pozo),
            "POZO_ID": _normalize_cell(pozo_id),
            "CAMPO": campo,
            "BLOQUE": bloque,
            "RIG": _normalize_cell(vRig),
            "NEGOCIO": _normalize_cell(vNegocio),
            "MOTIVO_PARADA": _normalize_cell(vMotivoParada),
            "RAZON_PULL": _normalize_cell(vRazonPull),
            "CIUDAD": _normalize_cell(vCiudad),
            "NUM_PULL": _parse_long(vNumPull),
            "FECHA_ARRANQUE": _combine_excel_date_time(vL2, vQ2),
            "FECHA_PARADA": _date_only_value(vL3),
            "FECHA_INICIO": _combine_excel_date_time(vL4, vQ4),
            "FECHA_FIN": _combine_excel_date_time(vL5, vQ5),
            "TIEMPO_FUN": _parse_long(vTiempoEqui),
            "TECNICOS": tecnicos,
            "NO_INSTALACION": selected_no_inst_payload,
        }
        existe = sb.table("CLIENTE_PULL") \
            .select("ID") \
            .eq("ID", id_base) \
            .execute()

        if existe.data:
            return jsonify({
                "ok": False,
                "tipo": "duplicado",
                "mensaje": "No se importó el reporte de pulling porque ya existe en la base de datos.",
                "duplicados": [id_base]
            }), 409

        # ✅ Insertar solo si no existe
        ok_ci, err_ci = _safe_write_with_variants(
            sb,
            "CLIENTE_PULL",
            cliente_payload,
            mode="insert"
        )

        ok_ci, err_ci = _safe_write_with_variants(sb, "CLIENTE_PULL", cliente_payload, mode="insert", on_conflict="ID")
        inserted["CLIENTE_PULL"] = 1 if ok_ci else 0
        if not ok_ci:
            _append_warning(warnings, "CLIENTE_PULL", "Falló el guardado en Supabase", payload=cliente_payload, detalle=err_ci)
        else:
            debug_steps.append({"tabla": "CLIENTE_PULL", "estado": "ok", "payload": cliente_payload})

        # ══════════════════════════════════════════════
        # 2) POZO_PULL
        # ══════════════════════════════════════════════
        vCabezal = _cell_value(ws, "D12")
        vConector = _cell_value(ws, "K12")
        vPesoSarta = _cell_value(ws, "Q12")
        vTuberiaPeso = _cell_value(ws, "W12")
        vNumTubos = _cell_value(ws, "AD12")
        vCsgPeso = _cell_value(ws, "AL12")
        vLinerPeso = _cell_value(ws, "E13")
        vDesv = _cell_value(ws, "O13")
        vYac = _cell_value(ws, "V13")
        vTemFondo = _cell_value(ws, "AB13")
        vProfSucc = _cell_value(ws, "AG13")
        vLongEqui = _cell_value(ws, "AQ13")

        pozo_payload = {
            "ID": id_base,
            "POZO_ID": _normalize_cell(pozo_id),
            "NO_INSTALACION": selected_no_inst_payload,
            "CABEZAL": _normalize_cell(vCabezal),
            "CONECTOR": _normalize_cell(vConector),
            "PESO_SARTA": _normalize_cell(vPesoSarta),
            "TUBERIA_PESO": _normalize_cell(vTuberiaPeso),
            "NUM_TUBOS": _normalize_cell(vNumTubos),
            "CSG_PESO": _normalize_cell(vCsgPeso),
            "LINER_PESO": _normalize_cell(vLinerPeso),
            "DESV": _normalize_cell(vDesv),
            "YACIMIENTO": _normalize_cell(vYac),
            "TEM_FONDO": _normalize_cell(vTemFondo),
            "PROF_SUCCION": _normalize_cell(vProfSucc),
            "LONG_EQUIPO": _normalize_cell(vLongEqui),
        }
        ok_pozo, err_pozo = _safe_write_with_variants(sb, "POZO_PULL", pozo_payload, mode="upsert", on_conflict="ID")
        inserted["POZO_PULL"] = 1 if ok_pozo else 0
        if not ok_pozo:
            _append_warning(warnings, "POZO_PULL", "Falló el guardado en Supabase", payload=pozo_payload, detalle=err_pozo)
        else:
            debug_steps.append({"tabla": "POZO_PULL", "estado": "ok", "payload": pozo_payload})

        # ══════════════════════════════════════════════
        # 3) SPOOLER_PULL
        # ══════════════════════════════════════════════
        spooler_payload = {
            "ID": id_base,
            "POZO_ID": _normalize_cell(pozo_id),
            "NO_INSTALACION": selected_no_inst_payload,
            "SPOOLER": _normalize_cell(vSpooler),
            "FECHA_SUBIDA_CABLE": _to_supabase_value(vFechaCable),
            "FECHA_SUBIDA_CAPILAR": _to_supabase_value(vFechaCap),
            "FECHA_BAJADA_CABLE": _to_supabase_value(vFechaCableEnd),
            "FECHA_BAJADA_CAPILAR": _to_supabase_value(vFechaCapEnd),
            "HORAS_SPOOLER_CABLE": _parse_float(vHorasCable),
            "HORAS_SPOOLER_CAPILAR": _parse_float(vHorasCap),
        }
        ok_sp, err_sp = _safe_write_with_variants(sb, "SPOOLER_PULL", spooler_payload, mode="upsert", on_conflict="ID")
        inserted["SPOOLER_PULL"] = 1 if ok_sp else 0
        if not ok_sp:
            _append_warning(warnings, "SPOOLER_PULL", "Falló el guardado en Supabase", payload=spooler_payload, detalle=err_sp)
        else:
            debug_steps.append({"tabla": "SPOOLER_PULL", "estado": "ok", "payload": spooler_payload})

        last_row_sheet = ws.max_row

        # ══════════════════════════════════════════════
        # 4) EQUIPO_PULL
        # ══════════════════════════════════════════════
        hdrPRow = 17
        lastColP = max(ws.max_column, column_index_from_string("AQ"))

        colPozoIdP = 0
        colEquipoP = 0
        colNoSerieP = 0
        colSerieP = 0
        colDescP = 0
        colEjeGiroP = 0
        colEjeRotoP = 0
        colHallCabezaP = 0
        colHallBaseP = 0
        colHousingArrP = 0
        colHousingSolP = 0
        colHousingCorrP = 0
        colHousingLimpP = 0

        for colIdxP in range(1, lastColP + 1):
            hTop = _clean_header(_merged_top_left_value(ws, 16, colIdxP))
            hSub = _clean_header(_merged_top_left_value(ws, 17, colIdxP))
            head = _clean_header(f"{hTop} {hSub}")

            if "COMPONENTE" in head:
                colEquipoP = colIdxP
            if "NUMERO_DE_SERIE" in head or ("NUMERO" in head and "SERIE" in head):
                colNoSerieP = colIdxP
            if "SERIE" in head and "NUMERO" not in head:
                colSerieP = colIdxP
            if "DESCRIPCION" in head:
                colDescP = colIdxP
            if "CONDICION_DEL_EJE" in head and "GIRO" in head:
                colEjeGiroP = colIdxP
            if "CONDICION_DEL_EJE" in head and "ROTO" in head:
                colEjeRotoP = colIdxP
            if "HALLAZGOS" in head and "CABEZA" in head:
                colHallCabezaP = colIdxP
            if "HALLAZGOS" in head and "BASE" in head:
                colHallBaseP = colIdxP
            if "CONDICION_DEL_HOUSING" in head and "ARRASTRE" in head:
                colHousingArrP = colIdxP
            if "CONDICION_DEL_HOUSING" in head and "SOLIDOS" in head:
                colHousingSolP = colIdxP
            if "CONDICION_DEL_HOUSING" in head and "CORROSION" in head:
                colHousingCorrP = colIdxP
            if "CONDICION_DEL_HOUSING" in head and "LIMPIO" in head:
                colHousingLimpP = colIdxP

        colHousingSolP = column_index_from_string("AK")
        colHousingCorrP = column_index_from_string("AN")
        colHousingLimpP = column_index_from_string("AQ")

        if colEquipoP == 0 and colNoSerieP == 0 and colSerieP == 0 and colDescP == 0 and            colEjeGiroP == 0 and colEjeRotoP == 0 and colHallCabezaP == 0 and colHallBaseP == 0 and            colHousingArrP == 0 and colHousingSolP == 0 and colHousingCorrP == 0 and colHousingLimpP == 0:
            _append_warning(warnings, "EQUIPO_PULL", "No se encontraron encabezados válidos para EQUIPO_PULL", fila=hdrPRow)
        else:
            regsP = 0
            for rP in range(hdrPRow + 1, last_row_sheet + 1):
                colsPList = [c for c in [colPozoIdP, colEquipoP, colNoSerieP, colSerieP, colDescP, colEjeGiroP, colEjeRotoP, colHallCabezaP, colHallBaseP, colHousingArrP, colHousingSolP, colHousingCorrP, colHousingLimpP] if c != 0]
                if colsPList:
                    tops = [_merged_top_row(ws, rP, c) for c in colsPList]
                    if len(set(tops)) == 1 and tops[0] != rP:
                        continue

                vPozoId = _normalize_cell(_merged_top_left_value(ws, rP, colPozoIdP)) if colPozoIdP else None
                vEquipo = _normalize_cell(_merged_top_left_value(ws, rP, colEquipoP)) if colEquipoP else None
                vNoSerie = _normalize_cell(_merged_top_left_value(ws, rP, colNoSerieP)) if colNoSerieP else None
                vSerie = _normalize_cell(_merged_top_left_value(ws, rP, colSerieP)) if colSerieP else None
                vDesc = _normalize_cell(_merged_top_left_value(ws, rP, colDescP)) if colDescP else None
                vEjeGiro = _normalize_cell(_merged_top_left_value(ws, rP, colEjeGiroP)) if colEjeGiroP else None
                vEjeRoto = _normalize_cell(_merged_top_left_value(ws, rP, colEjeRotoP)) if colEjeRotoP else None
                vHallCab = _normalize_cell(_merged_top_left_value(ws, rP, colHallCabezaP)) if colHallCabezaP else None
                vHallBase = _normalize_cell(_merged_top_left_value(ws, rP, colHallBaseP)) if colHallBaseP else None
                vHArr = _normalize_cell(_merged_top_left_value(ws, rP, colHousingArrP)) if colHousingArrP else None
                vHSol = _normalize_cell(_merged_top_left_value(ws, rP, colHousingSolP)) if colHousingSolP else None
                vHCorr = _normalize_cell(_merged_top_left_value(ws, rP, colHousingCorrP)) if colHousingCorrP else None
                vHLimp = _normalize_cell(_merged_top_left_value(ws, rP, colHousingLimpP)) if colHousingLimpP else None

                if all(v in (None, "") for v in [vPozoId, vEquipo, vNoSerie, vSerie, vDesc, vEjeGiro, vEjeRoto, vHallCab, vHallBase, vHArr, vHSol, vHCorr, vHLimp]):
                    break

                payload = {
                    "ID": f"{(vPozoId or pozo_id)}-{selected_no_inst}",
                    "POZO_ID": _normalize_cell(vPozoId or pozo_id),
                    "NO_INSTALACION": selected_no_inst_payload,
                    "EQUIPO": vEquipo,
                    "NO_SERIE": vNoSerie,
                    "SERIE": vSerie,
                    "DESCRIPCION": vDesc,
                    "EJE_GIRO": vEjeGiro,
                    "EJE_ROTO": vEjeRoto,
                    "HALLAZGOS_CABEZA": vHallCab,
                    "HALLAZGOS_BASE": vHallBase,
                    "HOUSING_ARRASTRE": vHArr,
                    "HOUSING_SOLIDOS": vHSol,
                    "HOUSING_CORROSION": vHCorr,
                    "HOUSING_LIMPIO": vHLimp,
                }
                ok, err = _safe_insert(sb, "EQUIPO_PULL", payload)
                inserted["EQUIPO_PULL"] = inserted.get("EQUIPO_PULL", 0) + (1 if ok else 0)
                if not ok:
                    warnings.append({"tabla": "EQUIPO_PULL", "fila": rP, "error": err})
                else:
                    regsP += 1
                    debug_steps.append({"tabla": "EQUIPO_PULL", "fila": rP, "estado": "ok"})

        # ══════════════════════════════════════════════
        # 5) CABLE_PULL
        # ══════════════════════════════════════════════
        hdrCRow = 35
        lastColC = ws.max_column

        colCable = 0
        colProp = 0
        colNoSerie = 0
        colCarreto = 0
        colDesc = 0
        colCant = 0
        colDanoF = 0
        colDanoE = 0
        colCorr = 0
        colSelloInt = 0
        colSello10 = 0
        colSelloExt = 0

        for colIdxC in range(1, lastColC + 1):
            h = _clean_header(_merged_top_left_value(ws, hdrCRow, colIdxC))
            if not h:
                continue
            if "CABLE" in h:
                colCable = colIdxC
            if "PROPIEDAD" in h:
                colProp = colIdxC
            if "NUMERO" in h and "SERIE" in h:
                colNoSerie = colIdxC
            if "CARRETO" in h:
                colCarreto = colIdxC
            if "DESCRIPCION" in h:
                colDesc = colIdxC
            if "CANTIDAD" in h:
                colCant = colIdxC
            if "DANO" in h and "FIS" in h:
                colDanoF = colIdxC
            if "DANO" in h and ("ELEC" in h or "ELECT" in h):
                colDanoE = colIdxC
            if "CORROS" in h:
                colCorr = colIdxC
            if "SELLO" in h and "INT" in h:
                colSelloInt = colIdxC
            if "SELLO" in h and "10" in h:
                colSello10 = colIdxC
            if "SELLO" in h and "EXT" in h:
                colSelloExt = colIdxC

        colSelloInt = column_index_from_string("AK")
        colSello10 = column_index_from_string("AN")
        colSelloExt = column_index_from_string("AQ")

        if colCable == 0 and colNoSerie == 0 and colDesc == 0:
            _append_warning(warnings, "CABLE_PULL", "No se detectaron encabezados para CABLE_PULL", fila=hdrCRow)
        else:
            regsC = 0
            for rC in range(hdrCRow + 1, last_row_sheet + 1):
                vCable = _normalize_cell(_merged_top_left_value(ws, rC, colCable)) if colCable else None
                vProp = _normalize_cell(_merged_top_left_value(ws, rC, colProp)) if colProp else None
                vNoS = _normalize_cell(_merged_top_left_value(ws, rC, colNoSerie)) if colNoSerie else None
                vCarr = _normalize_cell(_merged_top_left_value(ws, rC, colCarreto)) if colCarreto else None
                vDes = _normalize_cell(_merged_top_left_value(ws, rC, colDesc)) if colDesc else None
                vCant = _normalize_cell(_merged_top_left_value(ws, rC, colCant)) if colCant else None
                vDF = _normalize_cell(_merged_top_left_value(ws, rC, colDanoF)) if colDanoF else None
                vDE = _normalize_cell(_merged_top_left_value(ws, rC, colDanoE)) if colDanoE else None
                vCor = _normalize_cell(_merged_top_left_value(ws, rC, colCorr)) if colCorr else None
                vSI = _normalize_cell(_merged_top_left_value(ws, rC, colSelloInt)) if colSelloInt else None
                vS10 = _normalize_cell(_merged_top_left_value(ws, rC, colSello10)) if colSello10 else None
                vSE = _normalize_cell(_merged_top_left_value(ws, rC, colSelloExt)) if colSelloExt else None

                if all((val in (None, "") for val in [vCable, vNoS, vDes, vProp, vCarr, vCant])):
                    break

                payload = {
                    "ID": id_base,
                    "POZO_ID": _normalize_cell(pozo_id),
                    "NO_INSTALACION": selected_no_inst_payload,
                    "CABLE": vCable,
                    "PROPIEDAD": vProp,
                    "NO_SERIE": vNoS,
                    "CARRETO": vCarr,
                    "DESCRIPCION": vDes,
                    "CANTIDAD": vCant,
                    "DAÑO_FISI": vDF,
                    "DAÑO_ELEC": vDE,
                    "CORROSION": vCor,
                    "NO_SELLOINT": vSI,
                    "NO_SELLO10": vS10,
                    "NO_SELLOEXT": vSE,
                }
                ok, err = _safe_insert(sb, "CABLE_PULL", payload)
                inserted["CABLE_PULL"] = inserted.get("CABLE_PULL", 0) + (1 if ok else 0)
                if not ok:
                    warnings.append({"tabla": "CABLE_PULL", "fila": rC, "error": err})
                else:
                    regsC += 1
                    debug_steps.append({"tabla": "CABLE_PULL", "fila": rC, "estado": "ok"})

        # ══════════════════════════════════════════════
        # 6) ACCESORIOS_PULL
        # ══════════════════════════════════════════════
        hdrERow = 60
        lastColE = ws.max_column
        colEquipoE = column_index_from_string("A")
        colDescE = 0
        colCantE = 0

        for colIdxE in range(1, lastColE + 1):
            h = _clean_header(_merged_top_left_value(ws, hdrERow, colIdxE))
            if "DESCRIPCION" in h:
                colDescE = colIdxE
            if "CANTIDAD" in h:
                colCantE = colIdxE

        if colEquipoE == 0 or colDescE == 0 or colCantE == 0:
            _append_warning(warnings, "ACCESORIOS_PULL", "No se detectaron encabezados EQUIPO / DESCRIPCIÓN / CANTIDAD", fila=hdrERow)
        else:
            regsE = 0
            for rE in range(hdrERow + 1, last_row_sheet + 1):
                vEquipo = _normalize_cell(_merged_top_left_value(ws, rE, colEquipoE))
                vDesc = _normalize_cell(_merged_top_left_value(ws, rE, colDescE))
                vCant = _normalize_cell(_merged_top_left_value(ws, rE, colCantE))

                if (vEquipo in (None, "")) and (vDesc in (None, "")):
                    break

                descEN = []
                for c in range(5, 15):
                    cellVal = _normalize_cell(_merged_top_left_value(ws, rE, c))
                    descEN.append(str(cellVal) if cellVal not in (None, "") else "-")
                payload = {
                    "ID": id_base,
                    "POZO_ID": _normalize_cell(pozo_id),
                    "NO_INSTALACION": selected_no_inst_payload,
                    "EQUIPO": vEquipo,
                    "DESCRIPCION": " ".join(descEN).strip(),
                    "CANTIDAD": _parse_float(vCant) if _parse_float(vCant) is not None else (_parse_long(vCant)),
                }
                ok, err = _safe_insert(sb, "ACCESORIOS_PULL", payload)
                inserted["ACCESORIOS_PULL"] = inserted.get("ACCESORIOS_PULL", 0) + (1 if ok else 0)
                if not ok:
                    warnings.append({"tabla": "ACCESORIOS_PULL", "fila": rE, "error": err})
                else:
                    regsE += 1
                    debug_steps.append({"tabla": "ACCESORIOS_PULL", "fila": rE, "estado": "ok"})

        # ══════════════════════════════════════════════
        # 7) SUPERFICIE_PULL
        # ══════════════════════════════════════════════
        hdrSRow = 60
        lastColS = ws.max_column
        colEquipoS = column_index_from_string("AD")
        colNoSerieS = 0
        colKvaS = 0
        colPropS = 0

        for colIdxS in range(1, lastColS + 1):
            h = _clean_header(_merged_top_left_value(ws, hdrSRow, colIdxS))
            if "DE_SERIE" in h or ("NUMERO" in h and "SERIE" in h):
                colNoSerieS = colIdxS
            if "KVA" in h:
                colKvaS = colIdxS
            if "PROPIEDAD" in h:
                colPropS = colIdxS

        if colEquipoS == 0 or colNoSerieS == 0 or colKvaS == 0 or colPropS == 0:
            _append_warning(
                warnings,
                "SUPERFICIE_PULL",
                "No se detectaron los encabezados de SUPERFICIE_PULL",
                fila=hdrSRow,
                EQUIPO=colEquipoS,
                SERIE=colNoSerieS,
                KVA=colKvaS,
                PROP=colPropS,
            )
        else:
            regsS = 0
            for rs2 in range(hdrSRow + 1, last_row_sheet + 1):
                vEq = _normalize_cell(_merged_top_left_value(ws, rs2, colEquipoS))
                vSer = _normalize_cell(_merged_top_left_value(ws, rs2, colNoSerieS))
                vKva = _normalize_cell(_merged_top_left_value(ws, rs2, colKvaS))
                vPr = _normalize_cell(_merged_top_left_value(ws, rs2, colPropS))

                if all(v in (None, "") for v in [vEq, vSer, vKva, vPr]):
                    break

                payload = {
                    "ID": id_base,
                    "POZO_ID": _normalize_cell(pozo_id),
                    "NO_INSTALACION": selected_no_inst_payload,
                    "EQUIPO": vEq,
                    "NO_SERIE": vSer,
                    "KVA": vKva,
                    "PROPIEDAD": vPr,
                }
                ok, err = _safe_insert(sb, "SUPERFICIE_PULL", payload)
                inserted["SUPERFICIE_PULL"] = inserted.get("SUPERFICIE_PULL", 0) + (1 if ok else 0)
                if not ok:
                    warnings.append({"tabla": "SUPERFICIE_PULL", "fila": rs2, "error": err})
                else:
                    regsS += 1
                    debug_steps.append({"tabla": "SUPERFICIE_PULL", "fila": rs2, "estado": "ok"})

        return jsonify({
            "ok": True,
            "message": "Registros en Pulling cargados correctamente",
            "inserted": inserted,
            "warnings": warnings,
            "debug_steps": debug_steps,
        })

    except Exception as e:
        logger.error("Error importando reporte pulling: %s", traceback.format_exc())
        return jsonify({"ok": False, "error": str(e)}), 500

# ══════════════════════════════════════════════
# LISTAR TABLAS
# Requiere la función RPC "get_public_tables" en Supabase.
# Ejecuta este SQL una vez en el SQL Editor de Supabase:
#
#   CREATE OR REPLACE FUNCTION get_public_tables()
#   RETURNS SETOF text LANGUAGE sql SECURITY DEFINER AS $$
#     SELECT table_name::text
#     FROM information_schema.tables
#     WHERE table_schema = 'public'
#     ORDER BY table_name;
#   $$;
# ══════════════════════════════════════════════

@app.route("/api/tablas", methods=["GET"])
def listar_tablas():
    try:
        sb  = get_supabase()
        res = sb.rpc("get_public_tables").execute()
        tablas = res.data if res.data else []
        return jsonify({"ok": True, "tablas": tablas})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# DATOS DE TABLA CON PAGINACIÓN
# ══════════════════════════════════════════════

@app.route("/api/tabla/datos/<tabla>", methods=["GET"])
def obtener_datos_tabla(tabla):
    try:
        pagina     = request.args.get("pagina",    1,  type=int)
        por_pagina = request.args.get("por_pagina", 50, type=int)

        offset  = (pagina - 1) * por_pagina
        rng_end = offset + por_pagina - 1

        sb = get_supabase()

        # Obtener datos con conteo total exacto
        res = (
            sb.table(tabla)
            .select("*", count="exact")
            .range(offset, rng_end)
            .execute()
        )

        rows            = serialize_rows(res.data)
        total_registros = res.count or 0
        total_paginas   = (total_registros + por_pagina - 1) // por_pagina

        # Inferir columnas del primer registro
        columnas = (
            [{"name": k, "type": "text"} for k in res.data[0].keys()]
            if res.data else []
        )

        return jsonify({
            "ok":              True,
            "tabla":           tabla,
            "columnas":        columnas,
            "datos":           rows,
            "total":           len(rows),
            "total_registros": total_registros,
            "pagina":          pagina,
            "por_pagina":      por_pagina,
            "total_paginas":   total_paginas,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# ACTUALIZAR REGISTRO
# ══════════════════════════════════════════════
def limpiar_vacios(obj):
    limpio = {}

    for k, v in obj.items():
        if v is None:
            limpio[k] = None
        elif isinstance(v, str) and v.strip().lower() in ["", "none", "null", "undefined", "nan"]:
            limpio[k] = None
        else:
            limpio[k] = v

    return limpio
@app.route("/api/tabla/actualizar/<tabla>", methods=["POST"])
def actualizar_registro(tabla):
    try:
        data = request.get_json()
        if not data:
            return jsonify({"ok": False, "error": "No se envió datos"}), 400

        where_clause = data.pop("_where", {})
        
        where_clause = limpiar_vacios(where_clause)
        if not where_clause:
            return jsonify({"ok": False, "error": "Se requiere clausula WHERE para actualizar"}), 400
        if not data:
            return jsonify({"ok": False, "error": "No hay columnas para actualizar"}), 400

        sb    = get_supabase()
        query = sb.table(tabla).update(data)
        for col, val in where_clause.items():
            query = query.eq(col, val)
        res = query.execute()

        return jsonify({
            "ok":             True,
            "message":        f"Se actualizaron {len(res.data)} registro(s)",
            "filas_afectadas": len(res.data),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# INSERTAR REGISTRO
# ══════════════════════════════════════════════

@app.route("/api/tabla/insertar/<tabla>", methods=["POST"])
def insertar_registro(tabla):
    try:
        data = request.get_json()
        if not data:
            return jsonify({"ok": False, "error": "No se envió datos"}), 400

        sb = get_supabase()
        sb.table(tabla).insert(data).execute()
        return jsonify({"ok": True, "message": "Registro insertado exitosamente"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# ELIMINAR REGISTRO
# ══════════════════════════════════════════════

@app.route("/api/tabla/eliminar/<tabla>", methods=["POST"])
def eliminar_registro(tabla):
    try:
        data = request.get_json()
        if not data:
            return jsonify({"ok": False, "error": "No se envió datos"}), 400

        where_clause = data.get("_where", {})
        where_clause = limpiar_vacios(where_clause)
        if not where_clause:
            return jsonify({"ok": False, "error": "Se requiere clausula WHERE para eliminar"}), 400

        sb    = get_supabase()
        query = sb.table(tabla).delete()
        for col, val in where_clause.items():
            query = query.eq(col, val)
        res = query.execute()

        return jsonify({
            "ok":             True,
            "message":        f"Se eliminaron {len(res.data)} registro(s)",
            "filas_afectadas": len(res.data),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# EXPORTAR TABLA
# ══════════════════════════════════════════════

@app.route("/api/tabla/exportar/<tabla>", methods=["GET"])
def exportar_tabla(tabla):
    try:
        sb   = get_supabase()
        res  = sb.table(tabla).select("*").limit(10000).execute()
        rows = serialize_rows(res.data)
        return jsonify({
            "ok":              True,
            "tabla":           tabla,
            "total_registros": len(rows),
            "datos":           rows,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# DEBUG COLUMNAS
# ══════════════════════════════════════════════

@app.route("/api/debug/columnas/<tabla>", methods=["GET"])
def debug_columnas(tabla):
    try:
        sb  = get_supabase()
        res = sb.table(tabla).select("*").limit(1).execute()
        columnas = (
            [{"column_name": k, "data_type": "text"} for k in res.data[0].keys()]
            if res.data else []
        )
        return jsonify({"ok": True, "table": tabla, "columns": columnas})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# LISTAR POZOS
# ══════════════════════════════════════════════

@app.route("/api/pozos", methods=["GET"])
def listar_pozos():
    buscar = request.args.get("buscar", "").strip()

    try:
        sb = get_supabase()

        query = (
            sb.table("CLIENTE_INSTALACION")
            .select("ID")
            .not_.is_("ID", "null")
            .order("ID")
        )

        if buscar:
            query = query.ilike("ID", f"%{buscar}%")

        res = query.execute()

        data = []
        seen = set()

        for row in res.data:
            id_pozo = row.get("ID")

            if not id_pozo or id_pozo in seen:
                continue

            seen.add(id_pozo)

            data.append({
                "id": id_pozo,
                "label": id_pozo
            })

        return jsonify({"ok": True, "data": data})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# GUARDAR STOP
# ══════════════════════════════════════════════

@app.route("/api/stop/guardar", methods=["POST"])
def guardar_stop():
    try:
        data = request.get_json() or {}
        id_base = data.get("id")

        if not id_base:
            return jsonify({"ok": False, "error": "Falta ID"}), 400

        sb = get_supabase()

        ci = (
            sb.table("CLIENTE_INSTALACION")
            .select("POZO_ID,NO_INSTALACION")
            .eq("ID", id_base)
            .limit(1)
            .execute()
        )

        if not ci.data:
            return jsonify({"ok": False, "error": "No existe ese ID en CLIENTE_INSTALACION"}), 404

        pozo_id = ci.data[0].get("POZO_ID")
        no_instalacion = ci.data[0].get("NO_INSTALACION")

        payload = {
            "ID": id_base,
            "POZO_ID": str(pozo_id),
            "NO_INSTALACION": int(no_instalacion),
            "STOP_DATE": data.get("stop_date"),
            "RAZON_STOP": data.get("razon_stop"),
            "START_DATE": data.get("start_date"),
            "PULL_COMENT": data.get("pull_coment"),
            "GENERAL": data.get("general"),
            "GENERAL_ESPECIFICO": data.get("general_especifico"),
            "ESPECIFICO": data.get("especifico"),
            "QAQC": data.get("qaqc"),
        }

        res_abierto = (
            sb.table("STATUS")
            .select("ID,STOP_DATE")
            .eq("ID", id_base)
            .is_("START_DATE", "null")
            .order("STOP_DATE", desc=True)
            .limit(1)
            .execute()
        )

        if res_abierto.data:
            registro = res_abierto.data[0]

            update_data = {
                "STOP_DATE": payload["STOP_DATE"],
                "RAZON_STOP": payload["RAZON_STOP"],
                "START_DATE": payload["START_DATE"],
                "PULL_COMENT": payload["PULL_COMENT"],
                "GENERAL": payload["GENERAL"],
                "GENERAL_ESPECIFICO": payload["GENERAL_ESPECIFICO"],
                "ESPECIFICO": payload["ESPECIFICO"],
                "QAQC": payload["QAQC"],
            }

            (
                sb.table("STATUS")
                .update(update_data)
                .eq("ID", id_base)
                .eq("STOP_DATE", registro["STOP_DATE"])
                .is_("START_DATE", "null")
                .execute()
            )

            return jsonify({"ok": True, "message": "STOP abierto actualizado correctamente."})

        sb.table("STATUS").insert(payload).execute()

        return jsonify({"ok": True, "message": "Nuevo registro STOP creado correctamente."})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ══════════════════════════════════════════════
# DETALLE STOP
# ══════════════════════════════════════════════
@app.route("/api/login", methods=["POST"])
def login_usuario():
    try:
        data = request.get_json()

        nombre = data.get("nombre")
        contrasena = data.get("contrasena")

        if not nombre or not contrasena:
            return jsonify({
                "ok": False,
                "error": "Ingrese nombre de usuario y contraseña"
            }), 400

        sb = get_supabase()

        res = (
            sb.table("usuarios")
            .select("*")
            .eq("nombre", nombre)
            .eq("contrasena", contrasena)
            .limit(1)
            .execute()
        )

        if not res.data:
            return jsonify({
                "ok": False,
                "error": "Credenciales incorrectas"
            }), 401

        return jsonify({
            "ok": True,
            "message": "Login correcto",
            "usuario": res.data[0]
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500
@app.route("/api/stop/detalle", methods=["GET"])
def detalle_stop():
    id_base = request.args.get("id")

    if not id_base:
        return jsonify({"ok": False, "error": "Falta ID"}), 400

    try:
        sb = get_supabase()

        res = (
            sb.table("STATUS")
            .select("STOP_DATE,RAZON_STOP,START_DATE,PULL_COMENT,GENERAL,GENERAL_ESPECIFICO,ESPECIFICO,QAQC")
            .eq("ID", id_base)
            .not_.is_("STOP_DATE", "null")
            .is_("START_DATE", "null")
            .order("STOP_DATE", desc=True)
            .limit(1)
            .execute()
        )

        s = res.data[0] if res.data else {}

        return jsonify({
            "ok": True,
            "data": {
                "stop_date": s.get("STOP_DATE"),
                "razon_stop": s.get("RAZON_STOP"),
                "start_date": s.get("START_DATE"),
                "pull_coment": s.get("PULL_COMENT"),
                "general": s.get("GENERAL"),
                "general_especifico": s.get("GENERAL_ESPECIFICO"),
                "especifico": s.get("ESPECIFICO"),
                "qaqc": s.get("QAQC"),
            },
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# FALLAS / RIFTS
# ══════════════════════════════════════════════

@app.route("/api/fallas/general", methods=["GET"])
def get_general():
    try:
        sb  = get_supabase()
        res = (
            sb.table("RIFTS_FAILURES")
            .select("GENERAL")
            .not_.is_("GENERAL", "null")
            .order("GENERAL")
            .execute()
        )
        seen, rows = set(), []
        for r in res.data:
            v = r.get("GENERAL")
            if v and v not in seen:
                seen.add(v); rows.append(v)
        return jsonify({"ok": True, "data": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/fallas/general_especifico", methods=["GET"])
def get_general_especifico():
    general = request.args.get("general")
    try:
        sb    = get_supabase()
        query = (
            sb.table("RIFTS_FAILURES")
            .select("GENERAL_ESPECIFICO")
            .not_.is_("GENERAL_ESPECIFICO", "null")
            .order("GENERAL_ESPECIFICO")
        )
        if general:
            query = query.eq("GENERAL", general)
        res = query.execute()
        seen, rows = set(), []
        for r in res.data:
            v = r.get("GENERAL_ESPECIFICO")
            if v and v not in seen:
                seen.add(v); rows.append(v)
        return jsonify({"ok": True, "data": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/fallas/especifico", methods=["GET"])
def get_especifico():
    general            = request.args.get("general")
    general_especifico = request.args.get("general_especifico")
    try:
        sb    = get_supabase()
        query = (
            sb.table("RIFTS_FAILURES")
            .select("ESPECIFICO")
            .not_.is_("ESPECIFICO", "null")
            .order("ESPECIFICO")
        )
        if general:
            query = query.eq("GENERAL", general)
        if general_especifico:
            query = query.eq("GENERAL_ESPECIFICO", general_especifico)
        res = query.execute()
        seen, rows = set(), []
        for r in res.data:
            v = r.get("ESPECIFICO")
            if v and v not in seen:
                seen.add(v); rows.append(v)
        return jsonify({"ok": True, "data": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════

@app.route("/api/dashboard/resumen", methods=["GET"])
def dashboard_resumen():
    try:
        sb = get_supabase()

        # Total pozos únicos
        res_total = (
            sb.table("CLIENTE_INSTALACION")
            .select("POZO_ID", count="exact")
            .not_.is_("POZO_ID", "null")
            .execute()
        )
        total_pozos = res_total.count or 0

        # Pozos STOP: tienen STOP_DATE pero NO START_DATE
        res_stop = (
            sb.table("STATUS")
            .select("POZO_ID", count="exact")
            .not_.is_("POZO_ID",   "null")
            .not_.is_("STOP_DATE", "null")
            .is_("START_DATE", "null")
            .execute()
        )
        total_stop = res_stop.count or 0

        # Pozos PULLED
        res_pulled = (
            sb.table("CLIENTE_PULL")
            .select("POZO_ID", count="exact")
            .not_.is_("POZO_ID", "null")
            .execute()
        )
        total_pulled  = res_pulled.count or 0
        total_running = max(0, total_pozos - total_stop - total_pulled)

        return jsonify({
            "ok": True,
            "data": {
                "total_pozos": total_pozos,
                "running":     total_running,
                "stop":        total_stop,
                "pulled":      total_pulled,
            },
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════
# DETALLE POZO
# ══════════════════════════════════════════════

@app.route("/api/pozos/detalle", methods=["GET"])
def detalle_pozo():
    id_base = request.args.get("id")

    if not id_base:
        return jsonify({"ok": False, "error": "Falta ID"}), 400

    try:
        sb = get_supabase()

        ci_res = (
            sb.table("CLIENTE_INSTALACION")
            .select("ID,POZO_ID,NO_INSTALACION,CLIENTE,BLOQUE,CAMPO,FECHA_ARRANQUE,TIPO_NEGOCIO")
            .eq("ID", id_base)
            .limit(1)
            .execute()
        )

        if not ci_res.data:
            return jsonify({"ok": False, "error": "No existe ese ID en CLIENTE_INSTALACION"}), 404

        cliente_info = ci_res.data[0]

        pozo_id = cliente_info.get("POZO_ID")
        no_instalacion = cliente_info.get("NO_INSTALACION")

        ip2_res = (
            sb.table("INFPOZO2_INSTALACION")
            .select("ZONA_PRODUCTORA_INICIAL,NO_WORKOVER")
            .eq("POZO_ID", str(pozo_id))
            .eq("NO_INSTALACION", int(no_instalacion))
            .limit(1)
            .execute()
        )
        infpozo2_info = ip2_res.data[0] if ip2_res.data else {}

        st_res = (
            sb.table("STATUS")
            .select("STOP_DATE,GENERAL,GENERAL_ESPECIFICO,ESPECIFICO,RAZON_STOP,PULL_COMENT,START_DATE")
            .eq("POZO_ID", str(pozo_id))
            .eq("NO_INSTALACION", int(no_instalacion))
            .not_.is_("STOP_DATE", "null")
            .order("STOP_DATE", desc=True)
            .limit(1)
            .execute()
        )
        status_info = st_res.data[0] if st_res.data else {}

        cp_res = (
            sb.table("CLIENTE_PULL")
            .select("FECHA_INICIO,TIEMPO_FUN,NUM_PULL,RAZON_PULL,FECHA_PARADA")
            .eq("POZO_ID", str(pozo_id))
            .eq("NO_INSTALACION", int(no_instalacion))
            .order("FECHA_INICIO", desc=True)
            .limit(1)
            .execute()
        )
        pull_info = cp_res.data[0] if cp_res.data else {}

        ef_res = (
            sb.table("EQUIPOFONDO_INSTALACION")
            .select("EQUIPO,NO_SERIE,DESCRIPCION,NO_PARTE,LONGITUD,PROPIEDAD")
            .eq("POZO_ID", str(pozo_id))
            .eq("NO_INSTALACION", int(no_instalacion))
            .order("EQUIPO")
            .execute()
        )

        es_res = (
            sb.table("EQUISUPERFICIE_INSTALACION")
            .select("EQUIPO,NO_SERIE,DESCRIPCION,CONDICION,PROPIEDAD")
            .eq("POZO_ID", str(pozo_id))
            .eq("NO_INSTALACION", int(no_instalacion))
            .order("EQUIPO")
            .execute()
        )

        estado = "RUNNING"
        if pull_info:
            estado = "PULLED"
        elif status_info:
            if status_info.get("PULL_COMENT"):
                estado = "PULLED"
            elif status_info.get("STOP_DATE") and not status_info.get("START_DATE"):
                estado = "STOP"
            else:
                estado = "RUNNING"

        return jsonify({
            "ok": True,
            "data": {
                "datos_pozo": {
                    "pozo_id": pozo_id,
                    "no_instalacion": no_instalacion,
                    "cliente": cliente_info.get("CLIENTE"),
                    "bloque": cliente_info.get("BLOQUE"),
                    "campo": cliente_info.get("CAMPO"),
                    "arena": infpozo2_info.get("ZONA_PRODUCTORA_INICIAL"),
                },
                "pozo": {
                    "status": estado,
                    "stop": pull_info.get("FECHA_PARADA"),
                    "contrato": cliente_info.get("TIPO_NEGOCIO"),
                    "run_life": pull_info.get("TIEMPO_FUN"),
                    "arranque": cliente_info.get("FECHA_ARRANQUE"),
                },
                "stop": {
                    "fecha_stop": status_info.get("STOP_DATE"),
                    "general": status_info.get("GENERAL"),
                    "general_especifico": status_info.get("GENERAL_ESPECIFICO"),
                    "especifico": status_info.get("ESPECIFICO"),
                    "comentario_stop": status_info.get("RAZON_STOP"),
                    "pull_coment": status_info.get("PULL_COMENT"),
                    "fecha_start": status_info.get("START_DATE"),
                },
                "pulling": {
                    "fecha_pulling": pull_info.get("FECHA_INICIO"),
                    "workover": infpozo2_info.get("NO_WORKOVER"),
                    "num_pulling": pull_info.get("NUM_PULL"),
                },
                "comentarios": {
                    "comentario_tecnico": status_info.get("PULL_COMENT"),
                    "comentario_reporte": pull_info.get("RAZON_PULL"),
                },
                "equipo_fondo": serialize_rows(ef_res.data),
                "equipo_superficie": serialize_rows(es_res.data),
            },
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500



# ══════════════════════════════════════════════
# EXPORTAR REPORTES EN EXCEL
# ══════════════════════════════════════════════

def _excel_safe_value(value):
    """Convierte valores a algo que openpyxl escriba sin romper formatos."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, dt_time):
        return value
    return value


def _safe_str(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date, dt_time)):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return str(value)


def _looks_like_pozo(value) -> bool:
    return value is not None and str(value).strip() != ""


def _normalize_pozo_key(value) -> str:
    return _safe_str(value).strip()


def _fetch_all_rows(sb, table_name: str, *, select: str = "*", order: list[tuple[str, bool]] | None = None, page_size: int = 1000):
    """
    Lee una tabla de Supabase por páginas para evitar el límite de filas por respuesta.
    """
    all_rows = []
    offset = 0

    while True:
        query = sb.table(table_name).select(select).range(offset, offset + page_size - 1)
        if order:
            for col, desc in order:
                query = query.order(col, desc=desc)
        res = query.execute()
        chunk = res.data or []
        all_rows.extend(chunk)
        if len(chunk) < page_size:
            break
        offset += page_size

    return serialize_rows(all_rows)


def _build_index_first(rows, key_field: str = "POZO_ID", *, predicate=None):
    idx = {}
    for row in rows:
        pozo = _normalize_pozo_key(row.get(key_field))
        if not _looks_like_pozo(pozo):
            continue
        if predicate is not None and not predicate(row):
            continue
        if pozo not in idx:
            idx[pozo] = row
    return idx


def _build_index_latest(rows, key_field: str = "POZO_ID", date_field: str | None = None, *, predicate=None):
    """
    Toma el último registro por POZO. Si date_field existe, prioriza el más reciente.
    """
    grouped = {}
    for row in rows:
        pozo = _normalize_pozo_key(row.get(key_field))
        if not _looks_like_pozo(pozo):
            continue
        if predicate is not None and not predicate(row):
            continue
        grouped.setdefault(pozo, []).append(row)

    out = {}
    for pozo, items in grouped.items():
        if date_field:
            def _sort_key(r):
                v = r.get(date_field)
                if isinstance(v, datetime):
                    return v
                if isinstance(v, date):
                    return datetime.combine(v, dt_time.min)
                if v is None:
                    return datetime.min
                s = str(v).strip()
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
                    try:
                        return datetime.strptime(s, fmt)
                    except Exception:
                        pass
                return datetime.min
            items = sorted(items, key=_sort_key, reverse=True)
        out[pozo] = items[0]
    return out


def _match_equipo(value, patterns):
    text = _safe_str(value).upper()
    return any(p.upper() in text for p in patterns)


def _first_nonempty(*values):
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def _as_date_or_none(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, dt_time.min)
    if not value:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def _extract_tipo_motor(txt):
    if txt in (None, ""):
        return None
    m = re.search(r"N\s*([0-9]+)\s*,\s*([^,]+?)\s*,", str(txt), flags=re.IGNORECASE)
    return m.group(2).strip() if m else None


def _split_tokens(text):
    s = _safe_str(text).strip()
    if not s:
        return []
    s = re.sub(r"\s+", " ", s)
    return s.split(" ")


def _extraer_etapas_lp(txt):
    s = _safe_str(txt)
    m = re.search(r"(\d+)\s*ETAP", s, flags=re.IGNORECASE)
    return _parse_long(m.group(1)) if m else None

# Compatibilidad: el resto del archivo llama a _extract_etapas_lp
_extract_etapas_lp = _extraer_etapas_lp


def _autosize_columns(ws, max_col: int):
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=c, max_col=c):
            v = row[0].value
            if v is None:
                continue
            text = _safe_str(v)
            max_len = max(max_len, min(len(text), 45))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 24))


def _merge_title(ws, row: int, start_col: int, end_col: int, title: str, fill_color: str = "0070C0", font_color: str = "FFFFFF"):
    if start_col > end_col:
        return
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col)
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(bold=True, color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_block_headers(ws, row: int, start_col: int, headers: list[str]):
    for idx, header in enumerate(headers, start=0):
        cell = ws.cell(row=row, column=start_col + idx)
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)



# =============================================================
# REPORTE PEC - Adaptado del VBA original
# =============================================================
# =============================================================

# =============================================================
# CONSTANTES DE ESTILO
# =============================================================
PEC_FONT_NAME    = "Univers 47 CondensedLight"
PEC_HEADER_FILL  = "000080"   # azul marino #000080 (RGB 0,0,128)
PEC_HEADER_FONT  = "FFFFFF"   # letra blanca para los encabezados
PEC_HEADER_ROW   = 3
PEC_DATA_START   = 4
PEC_TOTAL_COLS   = 128


# =============================================================
# HELPERS PORTADOS DEL VBA  (prefijo _pec_)
# =============================================================
def _pec_s(v):
    return "" if v is None else str(v)

def _pec_instr(start, hay, needle, ci=True):
    """InStr(start, hay, needle [, vbTextCompare]). 1-based, 0 si no hay."""
    if not hay or not needle:
        return 0
    if ci:
        i = hay.upper().find(needle.upper(), start - 1)
    else:
        i = hay.find(needle, start - 1)
    return i + 1 if i >= 0 else 0

def _pec_instr_rev(hay, needle):
    if not hay or not needle:
        return 0
    i = hay.rfind(needle)
    return i + 1 if i >= 0 else 0

def _pec_val(s):
    """Equivalente Val(): número inicial. Devuelve int/float, 0 si nada."""
    if s is None:
        return 0
    s = str(s).strip().replace("/", "")
    if not s:
        return 0
    m = re.match(r"[-+]?\d+(\.\d+)?", s)
    if not m:
        return 0
    try:
        n = float(m.group(0))
        return int(n) if n.is_integer() else n
    except ValueError:
        return 0


# ---------- Extracciones de motor ----------
def _pec_extraer_hp(txt):
    s = _pec_s(txt)
    if not s: return ""
    p = _pec_instr(1, s, "HP")
    if p == 0: return ""
    seg = s[:p-1]
    pc = _pec_instr_rev(seg, ",")
    if pc == 0: return ""
    return _pec_val(seg[pc:].strip())

def _pec_extraer_v(txt):
    s = _pec_s(txt)
    if not s: return ""
    p_hp = _pec_instr(1, s, "HP")
    if p_hp == 0: return ""
    p_v = _pec_instr(p_hp + 2, s, "V")
    if p_v == 0: return ""
    return _pec_val(s[p_hp+2:p_v-1].replace("/", "").strip())

def _pec_extraer_a(txt):
    s = _pec_s(txt)
    if not s: return ""
    p_hp = _pec_instr(1, s, "HP")
    if p_hp == 0: return ""
    p_v = _pec_instr(p_hp + 2, s, "V")
    if p_v == 0: return ""
    p_a = _pec_instr(p_v + 1, s, "A")
    if p_a == 0: return ""
    return s[p_v+1:p_a-1].replace("/", "").strip()

def _pec_extraer_tipo_motor(txt):
    s = _pec_s(txt)
    if not s: return ""
    p_n = _pec_instr(1, s, "N")
    if p_n == 0: return ""
    pc = _pec_instr(p_n, s, ",", ci=False)
    if pc == 0: return ""
    seg = s[p_n:pc-1].strip()
    val_str = str(_pec_val(seg))
    if "." in val_str:
        val_str = val_str.split(".")[0]
    return seg[len(val_str):]

def _pec_extraer_serie_diam_motor(txt):
    s = _pec_s(txt)
    if not s: return ""
    p_n = _pec_instr(1, s, "N")
    if p_n == 0: return ""
    pc = _pec_instr(p_n, s, ",", ci=False)
    if pc == 0: return ""
    v = _pec_val(s[p_n:pc-1].strip())
    return v if v != 0 else ""

def _pec_extraer_rpm(txt):
    s = _pec_s(txt)
    if not s: return ""
    tipo = _pec_extraer_tipo_motor(txt)
    if tipo and str(tipo).strip().upper() == "AM":
        return 3492
    p_rpm = _pec_instr(1, s, "RPM")
    if p_rpm == 0: return ""
    seg = s[:p_rpm-1]
    pc = _pec_instr_rev(seg, ",")
    if pc == 0: return ""
    return _pec_val(seg[pc:].strip())

def _pec_extraer_serie_diam_primera_n(texto):
    """Primer 'N' seguido de dígitos -> esos dígitos (string)."""
    if texto is None: return ""
    s = str(texto)
    n = len(s)
    for i in range(n - 1):
        if s[i].upper() == "N" and s[i+1].isdigit():
            j = i + 1
            dig = ""
            while j < n and s[j].isdigit():
                dig += s[j]
                j += 1
            if dig:
                return dig
    return ""


# ---------- Tokens LP / Etapas / Código por modelo ----------
def _pec_limpiar_inicio_lp(texto):
    s = (texto or "").strip()
    if s.upper().startswith("PUMP,"):
        s = s[5:].strip()
    return s

def _pec_primer_token_lp(texto):
    if texto is None: return ""
    s = _pec_limpiar_inicio_lp(str(texto))
    if not s: return ""
    coma = s.find(",")
    if coma >= 0:
        return s[:coma].strip()
    sp = s.find(" ")
    return s[:sp] if sp >= 0 else s

def _pec_segundo_token_lp(texto):
    if texto is None: return ""
    s = _pec_limpiar_inicio_lp(str(texto))
    if not s: return ""
    coma = s.find(",")
    if coma >= 0:
        resto = s[coma+1:].strip()
        if not resto: return ""
        coma2 = resto.find(",")
        if coma2 >= 0:
            return resto[:coma2].strip()
        sp = resto.find(" ")
        return resto[:sp] if sp >= 0 else resto
    sp = s.find(" ")
    if sp < 0: return ""
    resto = s[sp+1:].lstrip()
    if not resto: return ""
    sp2 = resto.find(" ")
    return resto[:sp2] if sp2 >= 0 else resto

def _pec_extraer_etapas_lp(texto):
    if texto is None: return ""
    s = str(texto).upper()
    p = s.find("STG")
    if p < 0: return ""
    i = p - 1
    while i >= 0 and s[i] == " ":
        i -= 1
    dig = ""
    while i >= 0 and s[i].isdigit():
        dig = s[i] + dig
        i -= 1
    if dig:
        try: return int(dig)
        except ValueError: return ""
    return ""

def _pec_codigo_segun_modelo_lp(texto):
    """Mapeo VBA: H->406, A->272, B->319, F->362, P->535."""
    if texto is None: return ""
    modelo = _pec_primer_token_lp(texto)
    if not modelo or len(modelo) < 2:
        return ""
    return {"H": 406, "A": 272, "B": 319, "F": 362, "P": 535}.get(modelo[1].upper(), "")


# ---------- KVA / Pulsos ----------
def _pec_extraer_kva(texto):
    s = "" if texto is None else str(texto)
    for v in (" KVA", " kva", " Kva", " kvA"):
        s = s.replace(v, "KVA")
    u = s.upper()
    p = u.find("KVA")
    if p < 0: return ""
    numero = ""
    i = p - 1
    while i >= 0 and u[i].isdigit():
        numero = u[i] + numero
        i -= 1
    return f"{numero}KVA" if numero else ""

def _pec_extraer_pulsos(texto):
    s = "" if texto is None else str(texto)
    u = s.upper()
    p = u.find("P")
    if p < 0: return ""
    numero = ""
    i = p - 1
    while i >= 0:
        ch = u[i]
        if ch.isdigit():
            numero = ch + numero
        elif ch == " ":
            if numero: break
        else:
            break
        i -= 1
    return f"{numero} PULSOS" if numero else ""


# ---------- Parsing TMP para Protector inferior/upper ----------
def _pec_parse_prt_sel_tmp(desc):
    """
    MODELO_PRT_SEL = primera palabra después de 'TMP ' (hasta espacio o coma)
    SERIE_DIAM_PRT_SEL = Val(Mid entre ' N' y 'TMP')
    """
    if desc is None: return "", ""
    s = str(desc)
    if not s: return "", ""
    p_tmp = _pec_instr(1, s, "TMP", ci=False)
    if p_tmp == 0: return "", ""
    resto = s[p_tmp - 1 + 3:]
    coma = resto.find(",")
    if coma >= 0:
        resto = resto[:coma]
    resto = resto.strip()
    sp = resto.find(" ")
    modelo = resto[:sp] if sp >= 0 else resto
    p_sn = _pec_instr(1, s, " N", ci=False)
    serie_diam = ""
    if p_sn > 0 and p_tmp > 0:
        v = _pec_val(s[p_sn+1:p_tmp-1])
        if v != 0:
            serie_diam = v
    return modelo, serie_diam


# ---------- Concat: conector superficie / accesorios protectores ----------
def _pec_unir_conector_superficie(es_list, pozo, campo):
    if pozo is None or campo not in ("DESCRIPCION", "NO_SERIE"):
        return ""
    pozo_t = str(pozo).strip()
    vistos = []
    for r in es_list:
        rp = r.get("POZO_ID")
        if rp is None or str(rp).strip() != pozo_t:
            continue
        equipo = (r.get("EQUIPO") or "").upper()
        if "CONECTOR DE SUPERFICIE" in equipo or "CONECTOR DE SUPERFICE" in equipo:
            v = r.get(campo)
            if v is not None:
                t = str(v).strip()
                if t and t not in vistos:
                    vistos.append(t)
    return " | ".join(vistos)


_PEC_PATRONES_PROTECTORES = (
    "PROTECTORES CABLE",
    "PROTECTOR BLAST JOINT CABLE",
    "PROTECTORES BLAST JOINT",
    "PROTECTORES CABLE / MID JOINT",
    "PROTECTORES DE CABLE",
    "PROTECTORES DE CABLE/EQUIPO",
    "PROTECTORES MLP",
)

def _pec_concat_accesorios_inst(acc_list, pozo, no_inst, campo):
    if pozo is None or campo not in ("DESCRIPCION", "CANTIDAD", "PROPIEDAD"):
        return ""
    pozo_t = str(pozo).strip()
    inst_t = "" if no_inst is None else str(no_inst).strip()
    partes = []
    for r in acc_list:
        rp = "" if r.get("POZO_ID") is None else str(r.get("POZO_ID")).strip()
        ri = "" if r.get("NO_INSTALACION") is None else str(r.get("NO_INSTALACION")).strip()
        if rp != pozo_t or ri != inst_t:
            continue
        acc_txt = (r.get("ACCESORIOS") or "").upper()
        if not any(pat in acc_txt for pat in _PEC_PATRONES_PROTECTORES):
            continue
        v = r.get(campo)
        if v is None: continue
        t = str(v).strip()
        if t: partes.append(t)
    return " | ".join(partes)


# ---------- Predicados de equipo ----------
def _pec_equipo_contiene(equipo, *patrones):
    if equipo is None: return False
    e = str(equipo).upper()
    return any(p.upper() in e for p in patrones)

def _pec_equipo_es_motor(equipo):
    """Lista exacta del VBA para 'IN' de MOTOR."""
    if equipo is None: return False
    return str(equipo).strip().upper() in {
        "MOTOR", "MOTOR TANDEM", "MOTOR TANDEM O SUP",
        "MOTOR UNICO/SUP", "MOTOR UNICO / SUP", "MOTOR UNICO / SUP.",
        "MOTOR UPPER",
    }

def _pec_max_in(rows, key):
    """Equivalente a Max([campo]) de SQL."""
    best = None
    for r in rows:
        v = r.get(key)
        if v is None: continue
        s = str(v)
        if best is None or s > best:
            best = s
    return best


# =============================================================
# _merge_title  (versión actualizada con Univers 47 CondensedLight)
# Si ya tenías esta función en tu appp.py, BÓRRALA y usa esta.
# =============================================================
def _merge_title(ws, row, c1, c2, title, fill, font_color=PEC_HEADER_FONT):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row,   end_column=c2)
    cell = ws.cell(row, c1)
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=PEC_FONT_NAME, bold=True, color=font_color)
    cell.alignment = Alignment(horizontal="center",
                               vertical="center",
                               wrap_text=True)


# =============================================================
# Llenado uniforme de bombas
# =============================================================
def _pec_llenar_bomba(ws, dict_pozos, ef_by_pozo, patron,
                      col_modelo, col_etapas, col_tipo,
                      col_serie_diam, col_serie):
    """
    Lee la primera EQUIPOFONDO_INSTALACION cuyo EQUIPO contiene `patron`
    y escribe en las 5 columnas pasadas (en el orden VBA del llamador).
    """
    for pozo, idx in dict_pozos.items():
        out_row = idx + (PEC_DATA_START - 1)
        equipos = ef_by_pozo.get(pozo, [])
        eq = next((r for r in equipos if _pec_equipo_contiene(r.get("EQUIPO"), patron)), {})
        if not eq:
            continue
        desc = eq.get("DESCRIPCION") or ""
        ws.cell(out_row, col_modelo).value     = _excel_safe_value(_pec_primer_token_lp(desc))
        ws.cell(out_row, col_etapas).value     = _excel_safe_value(_pec_extraer_etapas_lp(desc))
        ws.cell(out_row, col_tipo).value       = _excel_safe_value(_pec_segundo_token_lp(desc))
        ws.cell(out_row, col_serie_diam).value = _excel_safe_value(_pec_codigo_segun_modelo_lp(desc))
        ws.cell(out_row, col_serie).value      = _excel_safe_value(eq.get("NO_SERIE") or "")


# =============================================================
# Headers PEC (128 cols, orden VBA)
# =============================================================
PEC_HEADERS = {
    1: "Pozo", 2: "ZONA", 3: "CAMPO", 4: "LOCACION", 5: "Arena",
    6: "Inst. Total", 7: "Start Date", 8: "Stop Date", 9: "Pulling Date",
    10: "Year Start", 11: "Year Stop", 12: "Year Pull", 13: "Downtime (Días)",
    14: "Today", 15: "Run life", 16: "FAILURE Status", 17: "Tiempo en Pozo",
    18: "ESP STATUS", 19: "CLIENTE ", 20: "CONTRATO", 21: "MODALIDAD CONTRACTUAL",
    22: "Mfg.", 23: "Modelo Sensor", 24: "Serie Sensor",
    25: "HP Motor", 26: "V Motor", 27: "A Motor", 28: "Tipo Motor",
    29: "RPM Motor", 30: "Serie (diam) Motor", 31: "Serie Motor",
    32: "Modelo PRT/SEL", 33: "Serie (diam) PRT/SEL", 34: "Serie PRT/SEL",
    35: "Modelo PRT/SEL", 36: "Serie (diam) PRT/SEL", 37: "Serie PRT/SEL",
    38: "Modelo ITK/SG", 39: "Serie (diam) ", 40: "Serie ITK/SG",
    41: "Modelo GH", 42: "Serie (diam) GH", 43: "Serie GH", 44: "# Etapas GH",
    45: "Modelo LP", 46: "# Etapas LP", 47: "Tipo LP", 48: "Serie (diam) ", 49: "Serie LP",
    50: "Modelo MP", 51: "# Etapas MP", 52: "Serie (diam) ", 53: "Tipo MP", 54: "Serie MP",
    55: "Modelo MP", 56: "# Etapas MP", 57: "Tipo MP", 58: "Serie (diam) ", 59: "Serie MP",
    60: "Modelo MP", 61: "# Etapas MP", 62: "Tipo MP", 63: "Serie (diam) ", 64: "Serie MP",
    65: "Modelo UP", 66: "# Etapas UP", 67: "Tipo UP", 68: "Serie (diam) ", 69: "Serie UP",
    70: "Marca Conector", 71: "Upper Conector", 72: "Lower Conector", 73: "Mandreal",
    74: "Serie 1 Cable", 75: "Serie 2 Cable", 76: "Serie 3 Cable", 77: "PROPIEDAD",
    78: "Tipo 1 Cable", 79: "Tipo 2 Cable", 80: "Tipo 3 Cable", 81: "Tipo 4 Cable",
    82: "Long 1 Cable (Pies)", 83: "Long 2 Cable (Pies)",
    84: "Long 3 Cable (Pies)", 85: "Long 4 Cable (Pies)",
    86: "MANDREL DOSIFICADOR DE QUIMICOS", 87: "MANEJADOR DE SOLIDOS",
    88: "PROPIEDAD ACC", 89: "MARCA PTR", 90: "CANTIDAD PTR", 91: "PROPIEDAD PTR",
    92: "MARCA  SDT", 93: "SERIAL  SDT", 94: "NOM. CONT.  SDT", 95: "KVA  SDT",
    96: "PULSOS  SDT", 97: "PROPIEDAD  SDT",
    98: "MARCA  SHIFT", 99: "SERIAL  SHIFT", 100: "NOM. CONT.  SHIFT",
    101: "KVA  SHIFT", 102: "PULSOS  SHIFT", 103: "PROPIEDAD  SHIFT",
    104: "MARCA VSD", 105: "SERIAL VSD", 106: "NOM. CONT.  VSD",
    107: "KVA VSD", 108: "PULSOS VSD", 109: "PROPIEDAD VSD",
    110: "MARCA SUT", 111: "SERIAL SUT", 112: "NOM. CONT.  SUT",
    113: "KVA SUT", 114: "PROPIEDAD SUT",
    115: "Modo de Falla General", 116: "Modo de Falla Especifico",
    117: "Componente en Falla", 118: "Sub-Componente en Falla",
    119: "Mecanismo de Falla General ", 120: "Mecanismo de Falla Específico",
    121: "Causa de Falla General", 122: "Causa de Falla Especifica",
    123: "Comentarios Pull - Tear Down", 124: "Falla directa ",
    125: "Falla indirecta", 126: "Plan de accion de fallas directas ",
    127: "PROPIEDAD", 128: "CABLE RENTA",
}


# =============================================================
# FUNCIÓN PRINCIPAL: REPORTE PEC
# =============================================================
def _build_reporte_pec_workbook(sb):
    """
    REPORTE PEC (Comando74_Click) - escribe por POSICIÓN de columna,
    replicando exactamente el VBA original.

    Estilo:
      - Encabezados (filas 1, 2 y 3) en fondo #000080 con letra blanca.
      - Toda la hoja (encabezados + datos) en fuente Univers 47 CondensedLight.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE PEC"
    ws.sheet_view.showGridLines = False

    # ---------- 1) Cargar tablas ----------
    ci_rows  = _fetch_all_rows(sb, "CLIENTE_INSTALACION",        select="*")
    ip_rows  = _fetch_all_rows(sb, "INFPOZO2_INSTALACION",       select="*")
    ef_rows  = _fetch_all_rows(sb, "EQUIPOFONDO_INSTALACION",    select="*")
    es_rows  = _fetch_all_rows(sb, "EQUISUPERFICIE_INSTALACION", select="*")
    acc_rows = _fetch_all_rows(sb, "ACCESORIOS_INSTALACION",     select="*")
    st_rows  = _fetch_all_rows(sb, "STATUS",                     select="*")

    # ---------- 2) Lista de pozos en orden de aparición ----------
    pozo_order = []
    seen = set()
    for r in ci_rows:
        p = _normalize_pozo_key(r.get("POZO_ID"))
        if _looks_like_pozo(p) and p not in seen:
            seen.add(p); pozo_order.append(p)
    dict_pozos = {p: i + 1 for i, p in enumerate(pozo_order)}

    # ---------- 3) Índices por pozo ----------
    ip_by_pozo = {}
    for r in ip_rows:
        p = _normalize_pozo_key(r.get("POZO_ID"))
        if p and p not in ip_by_pozo: ip_by_pozo[p] = r

    ef_by_pozo = {}
    for r in ef_rows:
        p = _normalize_pozo_key(r.get("POZO_ID"))
        if p: ef_by_pozo.setdefault(p, []).append(r)

    es_by_pozo = {}
    for r in es_rows:
        p = _normalize_pozo_key(r.get("POZO_ID"))
        if p: es_by_pozo.setdefault(p, []).append(r)

    st_by_pozo = {}
    for r in st_rows:
        p = _normalize_pozo_key(r.get("POZO_ID"))
        if p and p not in st_by_pozo: st_by_pozo[p] = r

    ci_by_pozo = {}
    for r in ci_rows:
        p = _normalize_pozo_key(r.get("POZO_ID"))
        if p and p not in ci_by_pozo: ci_by_pozo[p] = r

    # ---------- 4) Bandas de títulos (filas 1-2)  ----------
    _merge_title(ws, 1, 32,  37,  "PROTECTOR/SELLO",                         PEC_HEADER_FILL)
    _merge_title(ws, 1, 45,  69,  "BOMBAS",                                  PEC_HEADER_FILL)
    _merge_title(ws, 1, 92,  114, "EQUIPO ELÉCTRICO DE SUPERFICIE",          PEC_HEADER_FILL)
    _merge_title(ws, 1, 115, 122, "ESTRUCTURA DE DATOS DE EVENTOS DE FALLA", PEC_HEADER_FILL)

    sub_titles = [
        (23, 24,  "SENSOR"),
        (25, 31,  "MOTOR"),
        (32, 34,  "INFERIOR"),
        (35, 37,  "SUPERIOR"),
        (38, 40,  "INTAKE / SEP.GAS"),
        (41, 44,  "MANEJADOR DE GAS"),
        (45, 49,  "INFERIOR"),
        (50, 54,  "INTERMEDIA"),
        (55, 59,  "INTERMEDIA"),
        (60, 64,  "INTERMEDIA"),
        (65, 69,  "SUPERIOR"),
        (70, 73,  "CONECTOR"),
        (74, 84,  "CABLE"),
        (86, 88,  "ACCESORIOS"),
        (89, 91,  "PROTECTORES"),
        (92, 97,  "TRAFO. REDUCTOR (SDT)"),
        (98, 103, "TRAFO. SHIFT"),
        (104,109, "VARIADOR FRECUENCIA (VSD)"),
        (110,114, "TRANSFORMADOR  SALIDA ELEVADOR (SUT)"),
        (115,116, "MODO DE FALLA"),
        (117,118, "COMPONENTE / ITEM EN FALLA"),
        (119,120, "MECANISMO DE FALLA"),
        (121,122, "CAUSA DE FALLA"),
    ]
    for c1, c2, txt in sub_titles:
        _merge_title(ws, 2, c1, c2, txt, PEC_HEADER_FILL, PEC_HEADER_FONT)

    # Header (fila 3)
    for col, h in PEC_HEADERS.items():
        ws.cell(PEC_HEADER_ROW, col).value = h

    today = datetime.now()
    today_str = today.strftime("%d-%m-%Y")

    # ============================================================
    # 5) INFO POZO  -> cols 1, 4(CAMPO), 5(ZONA), 7, 10, 14, 19, 21
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        ci = ci_by_pozo.get(pozo, {})
        ip = ip_by_pozo.get(pozo, {})

        ws.cell(out_row, 1).value  = pozo
        ws.cell(out_row, 4).value  = _excel_safe_value(ci.get("CAMPO") or "")
        ws.cell(out_row, 5).value  = _excel_safe_value(ip.get("ZONA_PRODUCTORA_INICIAL") or "")
        ws.cell(out_row, 7).value  = _excel_safe_value(ci.get("FECHA_ARRANQUE") or "")

        fa = ci.get("FECHA_ARRANQUE")
        ys = ""
        if fa:
            try:
                ys = (fa.year if hasattr(fa, "year")
                      else datetime.fromisoformat(str(fa)[:10]).year)
            except Exception:
                m = re.search(r"\d{4}", str(fa))
                ys = int(m.group(0)) if m else ""
        ws.cell(out_row, 10).value = _excel_safe_value(ys)
        ws.cell(out_row, 14).value = today_str
        ws.cell(out_row, 19).value = _excel_safe_value(ci.get("CLIENTE") or "")
        ws.cell(out_row, 21).value = _excel_safe_value(ci.get("TIPO_NEGOCIO") or "")

    # ============================================================
    # 6) SENSOR + MOTOR  -> cols 23..31  (Max() en VBA)
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        equipos = ef_by_pozo.get(pozo, [])

        sensores = [r for r in equipos
                    if (r.get("EQUIPO") or "").strip().upper() == "SENSOR DE FONDO"]
        motores  = [r for r in equipos if _pec_equipo_es_motor(r.get("EQUIPO"))]

        desc_sensor    = _pec_max_in(sensores, "DESCRIPCION") or ""
        no_serie_sens  = _pec_max_in(sensores, "NO_SERIE")    or ""
        desc_motor     = _pec_max_in(motores,  "DESCRIPCION") or ""
        no_serie_motor = _pec_max_in(motores,  "NO_SERIE")    or ""

        ws.cell(out_row, 23).value = _excel_safe_value(desc_sensor)
        ws.cell(out_row, 24).value = _excel_safe_value(no_serie_sens)
        ws.cell(out_row, 25).value = _excel_safe_value(_pec_extraer_hp(desc_motor))
        ws.cell(out_row, 26).value = _excel_safe_value(_pec_extraer_v(desc_motor))
        ws.cell(out_row, 27).value = _excel_safe_value(_pec_extraer_a(desc_motor))
        ws.cell(out_row, 28).value = _excel_safe_value(_pec_extraer_tipo_motor(desc_motor))
        ws.cell(out_row, 29).value = _excel_safe_value(_pec_extraer_rpm(desc_motor))
        ws.cell(out_row, 30).value = _excel_safe_value(_pec_extraer_serie_diam_motor(desc_motor))
        ws.cell(out_row, 31).value = _excel_safe_value(no_serie_motor)

    # ============================================================
    # 7) PROTECTOR INFERIOR  -> cols 32..34
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        equipos = ef_by_pozo.get(pozo, [])
        prt = next((r for r in equipos
                    if _pec_equipo_contiene(r.get("EQUIPO"),
                                            "PROTECTOR", "PROT. INFERIOR",
                                            "PROTECTOR LOWER")), {})
        modelo, serie_diam = _pec_parse_prt_sel_tmp(prt.get("DESCRIPCION") or "")
        ws.cell(out_row, 32).value = _excel_safe_value(modelo)
        ws.cell(out_row, 33).value = _excel_safe_value(serie_diam)
        ws.cell(out_row, 34).value = _excel_safe_value(prt.get("NO_SERIE") or "")

    # ============================================================
    # 8) PROTECTOR UPPER  -> cols 35..37
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        equipos = ef_by_pozo.get(pozo, [])
        prt = next((r for r in equipos
                    if _pec_equipo_contiene(r.get("EQUIPO"),
                                            "PROT. TANDEM O SUP",
                                            "PROT. TANDEM O SUP.",
                                            "PROTECTOR UPPER")), {})
        modelo, serie_diam = _pec_parse_prt_sel_tmp(prt.get("DESCRIPCION") or "")
        ws.cell(out_row, 35).value = _excel_safe_value(modelo)
        ws.cell(out_row, 36).value = _excel_safe_value(serie_diam)
        ws.cell(out_row, 37).value = _excel_safe_value(prt.get("NO_SERIE") or "")

    # ============================================================
    # 9) GAS SEPARADOR / INTAKE  -> cols 38..40
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        equipos = ef_by_pozo.get(pozo, [])
        itk = next((r for r in equipos
                    if _pec_equipo_contiene(r.get("EQUIPO"), "GAS SEPARADOR / INTAKE")), {})
        desc = itk.get("DESCRIPCION") or ""
        ws.cell(out_row, 38).value = _excel_safe_value(desc)
        ws.cell(out_row, 39).value = _excel_safe_value(_pec_extraer_serie_diam_primera_n(desc))
        ws.cell(out_row, 40).value = _excel_safe_value(itk.get("NO_SERIE") or "")

    # cols 41..44 (Manejador de gas / GH) -> vacías

    # ============================================================
    # 10) BOMBA INFERIOR (45..49)
    # ============================================================
    _pec_llenar_bomba(ws, dict_pozos, ef_by_pozo, "BOMBA INFERIOR",
                      45, 46, 47, 48, 49)

    # ============================================================
    # 11) BOMBA MEDIA 1 (50..54)  -- ORDEN VBA distinto
    #     50 Modelo, 51 Etapas, 52 COD, 53 Tipo, 54 Serie
    # 12) BOMBA MEDIA 2 (55..59) -> 55 Mod, 56 Et, 57 Tipo, 58 COD, 59 Serie
    # 13) BOMBA MEDIA 3 (60..64) -> 60 Mod, 61 Et, 62 Tipo, 63 COD, 64 Serie
    # ============================================================
    _pec_llenar_bomba(ws, dict_pozos, ef_by_pozo, "BOMBA MEDIA 1",
                      50, 51, 53, 52, 54)  # OJO: 52=COD, 53=Tipo
    _pec_llenar_bomba(ws, dict_pozos, ef_by_pozo, "BOMBA MEDIA 2",
                      55, 56, 57, 58, 59)
    _pec_llenar_bomba(ws, dict_pozos, ef_by_pozo, "BOMBA MEDIA 3",
                      60, 61, 62, 63, 64)

    # ============================================================
    # 14) BOMBA SUPERIOR (65..69)
    # ============================================================
    _pec_llenar_bomba(ws, dict_pozos, ef_by_pozo, "BOMBA SUPERIOR",
                      65, 66, 67, 68, 69)

    # ============================================================
    # 15) CONECTOR DE SUPERFICIE  -> cols 70, 71
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        es = es_by_pozo.get(pozo, [])
        marca = _pec_unir_conector_superficie(es, pozo, "DESCRIPCION")
        upper = _pec_unir_conector_superficie(es, pozo, "NO_SERIE")
        if marca: ws.cell(out_row, 70).value = _excel_safe_value(marca)
        if upper: ws.cell(out_row, 71).value = _excel_safe_value(upper)

    # ============================================================
    # 16) CABLES  -> cols 74..85
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        equipos = ef_by_pozo.get(pozo, [])

        def _first_cable(pat):
            for r in equipos:
                if _pec_equipo_contiene(r.get("EQUIPO"), pat):
                    return r
            return {}

        c_inf  = _first_cable("CABLE INFERIOR")
        c_med1 = _first_cable("CABLE MEDIO 1")
        c_med2 = _first_cable("CABLE MEDIO 2")
        c_sup  = _first_cable("CABLE SUPERIOR")

        ws.cell(out_row, 74).value = _excel_safe_value(c_inf.get("NO_SERIE")  or "")
        ws.cell(out_row, 75).value = _excel_safe_value(c_med1.get("NO_SERIE") or "")
        ws.cell(out_row, 76).value = _excel_safe_value(c_sup.get("NO_SERIE")  or "")

        propiedad = " ".join([
            str(c_sup.get("PROPIEDAD")  or "").strip(),
            str(c_med1.get("PROPIEDAD") or "").strip(),
            str(c_inf.get("PROPIEDAD")  or "").strip(),
        ]).strip()
        ws.cell(out_row, 77).value = _excel_safe_value(propiedad)

        ws.cell(out_row, 78).value = _excel_safe_value(c_inf.get("DESCRIPCION")  or "")
        ws.cell(out_row, 79).value = _excel_safe_value(c_med1.get("DESCRIPCION") or "")
        ws.cell(out_row, 80).value = _excel_safe_value(c_med2.get("DESCRIPCION") or "")
        ws.cell(out_row, 81).value = _excel_safe_value(c_sup.get("DESCRIPCION")  or "")

        ws.cell(out_row, 82).value = _excel_safe_value(c_inf.get("LONGITUD")  or "")
        ws.cell(out_row, 83).value = _excel_safe_value(c_med1.get("LONGITUD") or "")
        ws.cell(out_row, 84).value = _excel_safe_value(c_sup.get("LONGITUD")  or "")
        ws.cell(out_row, 85).value = _excel_safe_value(c_med2.get("LONGITUD") or "")

    # ============================================================
    # 17) ACCESORIOS Y PROTECTORES  -> cols 86..91
    # ============================================================
    base_set = set()
    for r in ef_rows:
        base_set.add((_normalize_pozo_key(r.get("POZO_ID")) or "",
                      "" if r.get("NO_INSTALACION") is None else str(r.get("NO_INSTALACION"))))
    for r in acc_rows:
        base_set.add((_normalize_pozo_key(r.get("POZO_ID")) or "",
                      "" if r.get("NO_INSTALACION") is None else str(r.get("NO_INSTALACION"))))

    ya_escrito = set()
    for pozo, no_inst in base_set:
        if not pozo or pozo not in dict_pozos or pozo in ya_escrito:
            continue
        ya_escrito.add(pozo)
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        equipos = ef_by_pozo.get(pozo, [])

        def _first_eq(pat):
            for r in equipos:
                if str(r.get("NO_INSTALACION") or "") != no_inst:
                    continue
                if _pec_equipo_contiene(r.get("EQUIPO"), pat):
                    return r
            return {}

        mandrel    = _first_eq("MANDREL DOSIFICADOR")
        manejador  = _first_eq("MANEJADOR DE SÓLIDOS") or _first_eq("MANEJADOR DE SOLIDOS")

        propiedad_acc = " ".join([
            str(mandrel.get("PROPIEDAD")   or "").strip(),
            str(manejador.get("PROPIEDAD") or "").strip(),
        ]).strip()

        ws.cell(out_row, 86).value = _excel_safe_value(mandrel.get("DESCRIPCION")   or "")
        ws.cell(out_row, 87).value = _excel_safe_value(manejador.get("DESCRIPCION") or "")
        ws.cell(out_row, 88).value = _excel_safe_value(propiedad_acc)
        ws.cell(out_row, 89).value = _excel_safe_value(
            _pec_concat_accesorios_inst(acc_rows, pozo, no_inst, "DESCRIPCION"))
        ws.cell(out_row, 90).value = _excel_safe_value(
            _pec_concat_accesorios_inst(acc_rows, pozo, no_inst, "CANTIDAD"))
        ws.cell(out_row, 91).value = _excel_safe_value(
            _pec_concat_accesorios_inst(acc_rows, pozo, no_inst, "PROPIEDAD"))

    # ============================================================
    # 18) SDT - TRANSFORMADOR REDUCTOR  -> cols 93, 95, 97
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        es = es_by_pozo.get(pozo, [])
        sdt_list = sorted(
            [r for r in es if _pec_equipo_contiene(r.get("EQUIPO"), "TRANSFORMADOR REDUCTOR")],
            key=lambda r: ("" if r.get("NO_SERIE") is None else str(r.get("NO_SERIE")))
        )
        if not sdt_list: continue
        first = sdt_list[0]
        ws.cell(out_row, 93).value = _excel_safe_value(first.get("NO_SERIE")  or "")
        ws.cell(out_row, 95).value = _excel_safe_value(_pec_extraer_kva(first.get("DESCRIPCION")))
        ws.cell(out_row, 97).value = _excel_safe_value(first.get("PROPIEDAD") or "")

    # cols 98..103 (SHIFT) -> vacías

    # ============================================================
    # 19) VSD - VARIADOR DE FRECUENCIA  -> cols 105, 107, 108, 109
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        es = es_by_pozo.get(pozo, [])
        vsd_list = sorted(
            [r for r in es if _pec_equipo_contiene(r.get("EQUIPO"), "VARIADOR DE FRECUENCIA")],
            key=lambda r: ("" if r.get("NO_SERIE") is None else str(r.get("NO_SERIE")))
        )
        if not vsd_list: continue
        first = vsd_list[0]
        desc = first.get("DESCRIPCION") or ""
        ws.cell(out_row, 105).value = _excel_safe_value(first.get("NO_SERIE")  or "")
        ws.cell(out_row, 107).value = _excel_safe_value(_pec_extraer_kva(desc))
        ws.cell(out_row, 108).value = _excel_safe_value(_pec_extraer_pulsos(desc))
        ws.cell(out_row, 109).value = _excel_safe_value(first.get("PROPIEDAD") or "")

    # ============================================================
    # 20) SUT - TRANSFORMADOR ELEVADOR  -> cols 111, 113, 114
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        es = es_by_pozo.get(pozo, [])
        sut_list = sorted(
            [r for r in es if _pec_equipo_contiene(r.get("EQUIPO"), "TRANSFORMADOR ELEVADOR")],
            key=lambda r: ("" if r.get("NO_SERIE") is None else str(r.get("NO_SERIE")))
        )
        if not sut_list: continue
        first = sut_list[0]
        ws.cell(out_row, 111).value = _excel_safe_value(first.get("NO_SERIE")  or "")
        ws.cell(out_row, 113).value = _excel_safe_value(_pec_extraer_kva(first.get("DESCRIPCION")))
        ws.cell(out_row, 114).value = _excel_safe_value(first.get("PROPIEDAD") or "")

    # ============================================================
    # 21) STATUS  -> cols 115, 116
    # ============================================================
    for pozo in pozo_order:
        out_row = dict_pozos[pozo] + (PEC_DATA_START - 1)
        st = st_by_pozo.get(pozo, {})
        ws.cell(out_row, 115).value = _excel_safe_value(st.get("GENERAL")            or "")
        ws.cell(out_row, 116).value = _excel_safe_value(st.get("GENERAL_ESPECIFICO") or "")

    # cols 117..128 quedan vacías (el VBA tampoco las setea)

    # ============================================================
    # FORMATO FINAL
    #   - Filas 1, 2 y 3 (encabezados): fondo #000080, letra blanca, negrita.
    #   - Todas las celdas: fuente Univers 47 CondensedLight.
    # ============================================================
    last_row = max(PEC_HEADER_ROW, PEC_HEADER_ROW + len(pozo_order))
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A{PEC_HEADER_ROW}:{get_column_letter(PEC_TOTAL_COLS)}{last_row}"

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=PEC_HEADER_FILL)

    for r in range(1, last_row + 1):
        for c in range(1, PEC_TOTAL_COLS + 1):
            cell = ws.cell(r, c)
            cell.border = border

            if r <= PEC_HEADER_ROW:
                # Filas 1, 2 y 3 (encabezados)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center",
                                           wrap_text=True)
                cell.fill = header_fill
                cell.font = Font(name=PEC_FONT_NAME,
                                 bold=True,
                                 color=PEC_HEADER_FONT)
            else:
                # Filas de datos
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = Font(name=PEC_FONT_NAME, bold=False)
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")

    _autosize_columns(ws, PEC_TOTAL_COLS)
    return wb





def _build_reporte_general_workbook(sb):
    """
    REPORTE general (Comando45_Click) - versión Python.
    Genera una hoja dinámica por bloques, emulando la lógica de Access.
    """
    import re
    import unicodedata

    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE GENERAL"
    ws.sheet_view.showGridLines = False

    def _limpiar_texto_comparacion(valor):
        txt = _safe_str(valor).upper().strip()
        txt = unicodedata.normalize("NFKD", txt)
        txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
        txt = re.sub(r"\s+", " ", txt)
        return txt

    # Datos base
    ci_rows = _fetch_all_rows(sb, "CLIENTE_INSTALACION", select="*")
    d2_rows = _fetch_all_rows(sb, "DATA_2", select="*")
    inf2_rows = _fetch_all_rows(sb, "INFPOZO2_INSTALACION", select="*")
    status_rows = _fetch_all_rows(sb, "STATUS", select="*")
    pull_rows = _fetch_all_rows(sb, "CLIENTE_PULL", select="*")
    ef_rows = _fetch_all_rows(sb, "EQUIPOFONDO_INSTALACION", select="*")
    es_rows = _fetch_all_rows(sb, "EQUISUPERFICIE_INSTALACION", select="*")
    bienes_rows = _fetch_all_rows(sb, "BIENES", select="*")

    pozo_order, seen = [], set()
    for r in ci_rows:
        p = _normalize_pozo_key(r.get("POZO_ID"))
        if _looks_like_pozo(p) and p not in seen:
            seen.add(p)
            pozo_order.append(p)

    for r in d2_rows:
        p = _normalize_pozo_key(r.get("POZO"))
        if _looks_like_pozo(p) and p not in seen:
            seen.add(p)
            pozo_order.append(p)

    for r in inf2_rows:
        p = _normalize_pozo_key(r.get("POZO_ID"))
        if _looks_like_pozo(p) and p not in seen:
            seen.add(p)
            pozo_order.append(p)

    ci_by_pozo = _build_index_first(ci_rows)
    d2_by_pozo = _build_index_first([{"POZO_ID": r.get("POZO"), **r} for r in d2_rows])
    inf2_by_pozo = _build_index_first(inf2_rows)
    status_by_pozo = _build_index_latest(status_rows, date_field="STOP_DATE")
    pull_by_pozo = _build_index_latest(pull_rows, date_field="FECHA_INICIO")

    ef_by_pozo = {}
    for r in ef_rows:
        pozo = _normalize_pozo_key(r.get("POZO_ID"))
        if _looks_like_pozo(pozo):
            ef_by_pozo.setdefault(pozo, []).append(r)

    es_by_pozo = {}
    for r in es_rows:
        pozo = _normalize_pozo_key(r.get("POZO_ID"))
        if _looks_like_pozo(pozo):
            es_by_pozo.setdefault(pozo, []).append(r)

    bienes_by_pozo = {}
    for r in bienes_rows:
        pozo = _normalize_pozo_key(r.get("POZO"))
        if _looks_like_pozo(pozo):
            bienes_by_pozo.setdefault(pozo, []).append(r)

    def _buscar_pn_bienes(pozo, descripcion_equipo):
        pozo = _normalize_pozo_key(pozo)
        if not _looks_like_pozo(pozo) or not descripcion_equipo:
            return "", ""

        desc_equipo_limpia = _limpiar_texto_comparacion(descripcion_equipo)
        for r in bienes_by_pozo.get(pozo, []):
            desc_bienes_limpia = _limpiar_texto_comparacion(r.get("DESCRIPCION_CLIENTE", ""))
            if desc_bienes_limpia == desc_equipo_limpia:
                return _safe_str(r.get("PN_CLIENTE")), _safe_str(r.get("PN_SYTELINE"))

        return "", ""

    def _first_equipo(rows, patterns):
        for row in rows:
            if _match_equipo(row.get("EQUIPO"), patterns):
                return row
        return {}

    def _surface_rows(pozo):
        return es_by_pozo.get(pozo, [])

    def _fondo_rows(pozo):
        return ef_by_pozo.get(pozo, [])

    def _row_fondo(pozo, pat):
        row = dict(_first_equipo(_fondo_rows(pozo), [pat]) or {})
        descripcion = _safe_str(row.get("DESCRIPCION"))
        pn_cliente, pn_syteline = _buscar_pn_bienes(pozo, descripcion)
        row["PN_CLIENTE"] = pn_cliente
        row["PN_SYTELINE"] = pn_syteline
        return row

    # Block definitions
    blocks = []

    def add_block(block_type, title, headers, row_func):
        blocks.append({
            "type": block_type,
            "title": title,
            "headers": headers,
            "row_func": row_func,
        })

    add_block(
        "INFO",
        "INFORMACIÓN POZO",
        [
            "POZO_ID",
            "BLOQUE",
            "CAMPO",
            "NO_INSTALACION",
            "NO_WORKOVER",
            "ZONA_PRODUCTORA_INICIAL",
            "FECHA_ARRANQUE",
            "CLIENTE",
            "TIPO_NEGOCIO",
            "STATUS",
        ],
        lambda p: {
            **ci_by_pozo.get(p, {}),
            **inf2_by_pozo.get(p, {}),
            **d2_by_pozo.get(p, {}),
        }
    )

    add_block(
        "STATUS",
        "INFORMACIÓN POZO",
        ["STOP_DATE", "START_DATE", "RAZON_STOP"],
        lambda p: status_by_pozo.get(p, {})
    )

    fondo_defs = [
        ("NCDV", "NCDV"),
        ("SUBDESCARGA", "SUBDESCARGA"),
        ("DESCARGA", "DESCARGA"),
        ("BOMBA SUPERIOR", "BOMBA SUPERIOR"),
        ("BOMBA MEDIA 4", "BOMBA MEDIA 4"),
        ("BOMBA MEDIA 3", "BOMBA MEDIA 3"),
        ("BOMBA MEDIA 2", "BOMBA MEDIA 2"),
        ("BOMBA MEDIA 1", "BOMBA MEDIA 1"),
        ("BOMBA INFERIOR", "BOMBA INFERIOR"),
        ("BOMBA MULTIFASICA", "BOMBA MULTIFASICA"),
        ("GAS SEPARADOR / INTAKE", "GAS SEPARADOR / INTAKE"),
        ("PROTECTOR INFERIOR", "PROTECTOR INFERIOR"),
        ("PROTECTOR MEDIO", "PROTECTOR MEDIO"),
        ("PROTECTOR SUPERIOR", "PROTECTOR SUPERIOR"),
        ("MOTOR", "MOTOR"),
        ("SENSOR DE FONDO", "SENSOR DE FONDO"),
        ("MANDREL DOSIFICADOR", "MANDREL DOSIFICADOR"),
        ("CENTRALIZADOR", "CENTRALIZADOR"),
        ("MANEJADOR DE SÓLIDOS", "MANEJADOR DE SÓLIDOS"),
        ("CABLE SUPERIOR", "CABLE SUPERIOR"),
        ("CABLE MEDIO 2", "CABLE MEDIO 2"),
        ("CABLE MEDIO 1", "CABLE MEDIO 1"),
        ("CABLE INFERIOR", "CABLE INFERIOR"),
        ("CABLE EXTENSIÓN (MLE)", "CABLE EXTENSIÓN (MLE)"),
        ("CABLE TOTAL", "CABLE TOTAL"),
    ]

    for patterns_label, title in fondo_defs:
        add_block(
            "EQUIPO",
            title,
            [
                "DESCRIPCION",
                "NO_PARTE",
                "NO_SERIE",
                "PROPIEDAD",
                "LONGITUD",
                "PN_CLIENTE",
                "PN_SYTELINE",
            ],
            lambda p, pat=patterns_label: _row_fondo(p, pat)
        )

    surface_defs = [
        ("GENERACIÓN", "GENERACION", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("TRANSFORMADOR REDUCTOR", "TRANSFORMADOR REDUCTOR", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("VARIADOR DE FRECUENCIA", "VARIADOR DE FRECUENCIA", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("TRANSFORMADOR ELEVADOR", "TRANSFORMADOR ELEVADOR", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("PANEL DEL SENSOR", "PANEL DEL SENSOR", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("CHOKE", "CHOKE", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("SOPORTE CHOKE Y PANEL", "SOPORTE CHOKE Y PANEL", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("CONTROLADOR", "CONTROLADOR", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("SISTEMA DE MONITOREO", "SISTEMA DE MONITOREO", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
        ("CAJA DE VENTEO", "CAJA DE VENTEO", ["DESCRIPCION", "NO_SERIE", "PROPIEDAD"]),
    ]

    for pat, title, hdrs in surface_defs:
        add_block(
            "EQUIPO",
            title,
            hdrs,
            lambda p, pat=pat: _first_equipo(_surface_rows(p), [pat])
        )

    add_block(
        "EQUIPO",
        "CONECTOR SUPERFICIE",
        ["DESCRIPCION", "NO_SERIE_COMBINADO", "PROPIEDAD_COMBINADA"],
        lambda p: (lambda rows: {
            "DESCRIPCION": _first_nonempty(*[_safe_str(r.get("DESCRIPCION")) for r in rows]),
            "NO_SERIE_COMBINADO": " / ".join(
                [_safe_str(r.get("NO_SERIE")) for r in rows if _looks_like_pozo(r.get("NO_SERIE"))]
            ) or None,
            "PROPIEDAD_COMBINADA": " / ".join(
                [_safe_str(r.get("PROPIEDAD")) for r in rows if _looks_like_pozo(r.get("PROPIEDAD"))]
            ) or None,
        })([r for r in _surface_rows(p) if _match_equipo(r.get("EQUIPO"), ["CONECTOR DE SUPERFICIE", "CONECTOR SUPERFICIE"])])
    )

    add_block(
        "EQUIPO",
        "PULLING",
        ["PULL_COMENT_TXT"],
        lambda p: {"PULL_COMENT_TXT": status_by_pozo.get(p, {}).get("PULL_COMENT")}
    )

    add_block(
        "EQUIPO",
        "PULLING",
        ["NUM_PULL", "TIEMPO_FUN", "RAZON_PULL", "FECHA_PARADA", "TECNICOS"],
        lambda p: pull_by_pozo.get(p, {})
    )

    add_block(
        "STATUS",
        "PULLING",
        ["GENERAL", "GENERAL_ESPECIFICO", "ESPECIFICO", "QAQC"],
        lambda p: status_by_pozo.get(p, {})
    )

    # build headers/positions
    start_cols, end_cols, field_counts = [], [], []
    out_col = 1
    for b in blocks:
        start_cols.append(out_col)
        field_counts.append(len(b["headers"]))
        end_cols.append(out_col + len(b["headers"]) - 1)
        out_col = end_cols[-1] + 1
    total_cols = out_col - 1

    # rows 1-3
    fondo_start_idx = 2
    fondo_end_idx = fondo_start_idx + len(fondo_defs) - 1
    surface_start_idx = fondo_end_idx + 1
    surface_end_idx = surface_start_idx + len(surface_defs) - 1

    if len(blocks) > fondo_end_idx:
        _merge_title(
            ws,
            1,
            start_cols[fondo_start_idx],
            end_cols[fondo_end_idx],
            "EQUIPO FONDO",
            "0070C0"
        )

    if len(blocks) > surface_end_idx:
        _merge_title(
            ws,
            1,
            start_cols[surface_start_idx],
            end_cols[surface_end_idx],
            "EQUIPO SUPERFICIE",
            "0070C0"
        )

    _merge_title(ws, 2, start_cols[0], end_cols[1], "INFORMACIÓN POZO", "0070C0")
    for i in range(2, len(blocks)):
        _merge_title(ws, 2, start_cols[i], end_cols[i], blocks[i]["title"], "0070C0")

    # row 3 headers
    for i, block in enumerate(blocks):
        _write_block_headers(ws, 3, start_cols[i], block["headers"])

    # data rows
    for r_idx, pozo in enumerate(pozo_order, start=4):
        for i, block in enumerate(blocks):
            row = block["row_func"](pozo) or {}
            headers = block["headers"]

            for j, h in enumerate(headers):
                val = row.get(h)
                if block["type"] == "INFO" and h == "POZO_ID":
                    val = pozo
                ws.cell(r_idx, start_cols[i] + j).value = _excel_safe_value(val)

    last_row = max(3, 3 + len(pozo_order))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(total_cols)}{last_row}"

    # styling
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(1, last_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if r <= 3 else "left",
                vertical="center",
                wrap_text=True
            )

            if r == 1:
                cell.fill = PatternFill("solid", fgColor="0070C0")
                cell.font = Font(bold=True, color="FFFFFF")
            elif r == 2:
                cell.fill = PatternFill("solid", fgColor="0070C0")
                cell.font = Font(bold=True, color="FFFFFF")
            elif r == 3:
                cell.fill = PatternFill("solid", fgColor="DDEBF7")
                cell.font = Font(bold=True, color="000000")
            elif r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="DDEBF7")

    _autosize_columns(ws, total_cols)
    return wb


def _download_workbook_as_response(wb, filename: str):
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )


@app.route("/api/reportes/reporte", methods=["GET"])
def descargar_reporte_general():
    try:
        sb = get_supabase()
        wb = _build_reporte_general_workbook(sb)
        nombre = f"REPORTE_GENERAL_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _download_workbook_as_response(wb, nombre)
    except Exception as e:
        logger.error("Error generando REPORTE general: %s", traceback.format_exc())
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/reportes/reporte-pec", methods=["GET"])
def descargar_reporte_pec():
    try:
        sb = get_supabase()
        wb = _build_reporte_pec_workbook(sb)
        nombre = f"REPORTE_PEC_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _download_workbook_as_response(wb, nombre)
    except Exception as e:
        logger.error("Error generando REPORTE PEC: %s", traceback.format_exc())
        return jsonify({"ok": False, "error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)





